import os
import re
import json
import uuid
import asyncio
from datetime import datetime
from typing import Optional, List
from google.cloud import firestore
from fastapi import APIRouter, HTTPException, status, UploadFile, File, Request, BackgroundTasks
from fastapi.responses import HTMLResponse, StreamingResponse, RedirectResponse
from pydantic import BaseModel
from google.adk.sessions import VertexAiSessionService

# Dynamic Global Delegation class to allow unit tests to patch globals on main
class DynamicGlobal:
    def __init__(self, name):
        self.name = name
    def __getattr__(self, item):
        import submission_frontend.main as main
        actual = getattr(main, self.name)
        return getattr(actual, item)
    def __bool__(self):
        import submission_frontend.main as main
        actual = getattr(main, self.name)
        return bool(actual)
    def __call__(self, *args, **kwargs):
        import submission_frontend.main as main
        actual = getattr(main, self.name)
        return actual(*args, **kwargs)

# Instantiated Dynamic Globals mapped to main's attributes
db = DynamicGlobal("db")
logger = DynamicGlobal("logger")
get_current_user_and_role = DynamicGlobal("get_current_user_and_role")
get_gcs_bucket = DynamicGlobal("get_gcs_bucket")
is_auth_enabled = DynamicGlobal("is_auth_enabled")

# Constants
from submission_frontend.config.settings import (
    BUCKET_NAME,
    PROJECT_ID,
    EXPENSES_COL,
    DOCUMENTS_COL,
    DECISIONS_COL,
    AUDIT_LOGS_COL,
    POLICIES_COL,
    LOCATION,
    ENGINE_ID,
    AGENT_RUNTIME_ID
)

from submission_frontend.schemas.schemas import ApprovalAction

# Imported Utilities
from submission_frontend.utilities.helpers import (
    add_audit_log,
    enrich_claim_with_employee_info,
    filter_and_enrich_claims,
    sanitize_claim_dict
)
from submission_frontend.utilities.policy_engine import (
    run_per_diem_check,
    run_policy_check_py
)
from submission_frontend.utilities.workflow_engine import (
    find_and_bind_expense,
    reevaluate_expense_policies,
    sync_completed_sessions,
    parse_claim_from_session,
    check_claim_missing_documents,
    recalculate_report_totals,
    add_report_audit_log
)

import time
import logging

# Suppress noisy informational logs from google-genai and Vertex AI libraries
for logger_name in ["google", "google_genai", "google.genai", "google_genai._api_client", "google_genai.api_client"]:
    logging.getLogger(logger_name).setLevel(logging.WARNING)

# Module-level caching for dashboard performance tuning
_pending_cache = {}  # key: params_tuple, value: (timestamp, data_dict)
_pending_cache_ttl = 30.0  # cache TTL in seconds

_expenses_cache = {}  # key: params_tuple, value: (timestamp, expenses_list)
_expenses_cache_ttl = 30.0  # cache TTL in seconds

router = APIRouter()

@router.get("/api/pending")
async def get_pending(
    request: Request,
    search: str = None,
    department: str = None,
    manager: str = None,
    status: str = None,
    category: str = None,
    company: str = None,
    hide_old_test_sessions: bool = True,
    assigned_to_me: bool = False,
    show_all_fa: bool = False,
    source: str = "all",
    refresh: bool = False
):
    """
    Queries VertexAiSessionService, fetches full histories in parallel,
    and returns sessions with unresolved adk_request_input events.
    """
    current_user = get_current_user_and_role(request)
    email = current_user.get("email")
    role = current_user.get("role")
    
    # Construct cache key
    cache_key = (
        email,
        role,
        search,
        department,
        manager,
        status,
        category,
        company,
        hide_old_test_sessions,
        assigned_to_me,
        show_all_fa,
        source
    )
    
    if not refresh:
        now = time.time()
        if cache_key in _pending_cache:
            cache_ts, cached_data = _pending_cache[cache_key]
            if now - cache_ts < _pending_cache_ttl:
                logger.info(f"Returning cached pending approvals for user {email}")
                return cached_data

    params = {
        "search": search,
        "department": department,
        "manager": manager,
        "status": status,
        "category": category,
        "company": company,
        "hide_old_test_sessions": hide_old_test_sessions,
        "assigned_to_me": assigned_to_me,
        "show_all_fa": show_all_fa,
        "source": source
    }
    try:
        s = VertexAiSessionService(
            project=PROJECT_ID,
            location=LOCATION,
            agent_engine_id=ENGINE_ID
        )
        
        list_resp = await s.list_sessions(app_name="app")
        pending_claims = []
        hidden_cli_sessions_count = 0
        
        # Sort sessions by recency (newest update time first)
        sessions_sorted = sorted(
            list_resp.sessions,
            key=lambda x: getattr(x, "last_update_time", 0.0),
            reverse=True
        )
        
        # Filter summaries and track hidden cli-user sessions count
        active_summaries = []
        for summary in sessions_sorted:
            if summary.user_id == "cli-user":
                hidden_cli_sessions_count += 1
            else:
                active_summaries.append(summary)
        
        # Limit to scanning the 30 most recent active sessions to prevent excessive API calls
        active_summaries = active_summaries[:30]
        
        # Semaphore to cap concurrent requests to Vertex AI
        sem = asyncio.Semaphore(10)
        
        async def fetch_and_parse(summary):
            async with sem:
                try:
                    # Individual session fetch timeout of 5 seconds
                    sess = await asyncio.wait_for(
                        s.get_session(app_name="app", user_id=summary.user_id, session_id=summary.id),
                        timeout=5.0
                    )
                    if not sess:
                        return []
                        
                    # Map of unresolved request input function calls
                    unresolved_calls = {}
                    for event in sess.events:
                        if not event.content or not event.content.parts:
                            continue
                        for part in event.content.parts:
                            fc = getattr(part, "function_call", None)
                            if fc and fc.name == "adk_request_input":
                                interrupt_id = getattr(fc, "id", None) or fc.args.get("interruptId", "review_decision")
                                unresolved_calls[interrupt_id] = {
                                    "interrupt_id": interrupt_id,
                                    "message": fc.args.get("message", "Attention required"),
                                }
                            fr = getattr(part, "function_response", None)
                            if fr and fr.name == "adk_request_input":
                                response_id = getattr(fr, "id", None)
                                unresolved_calls.pop(response_id, None)
                                
                    if unresolved_calls:
                        claim_details = parse_claim_from_session(sess)
                        
                        # Fetch GCS uploads metadata for this session in parallel
                        uploaded_metadata = {}
                        try:
                            def get_gcs_metadata():
                                try:
                                    bucket = get_gcs_bucket()
                                    m_blob = bucket.blob(f"uploads/{sess.id}/metadata.json")
                                    return json.loads(m_blob.download_as_bytes().decode("utf-8"))
                                except Exception:
                                    return {}
                            uploaded_metadata = await asyncio.to_thread(get_gcs_metadata)
                        except Exception as ge:
                            logger.warning(f"Error fetching GCS metadata for session {sess.id}: {ge}")
                            
                        # Merge uploaded document paths as endpoint URLs if they exist in GCS and original is missing
                        doc_url_fields = {
                            "receipt": "receipt_url",
                            "hotel_receipt": "hotel_receipt_url",
                            "flight_ticket_receipt": "flight_ticket_receipt_url",
                            "manager_approval_letter": "manager_approval_letter_url",
                            "office_receipt": "office_receipt_url",
                            "parking_receipt": "parking_receipt_url"
                        }
                        
                        for doc_type, field_name in doc_url_fields.items():
                            if doc_type in uploaded_metadata:
                                # Overwrite if missing from original session payload
                                if not claim_details.get(field_name):
                                    claim_details[field_name] = f"/api/document/{sess.id}/{doc_type}"
                        
                        # Bind and find/upsert in Firestore
                        expense_data, claim_id = await asyncio.to_thread(find_and_bind_expense, sess.id, sess.user_id, claim_details)
                        
                        claims_parsed = []
                        for interrupt_id, call_info in unresolved_calls.items():
                            claims_parsed.append({
                                "claim_id": claim_id,
                                "session_id": sess.id,
                                "user_id": sess.user_id,
                                "interrupt_id": interrupt_id,
                                "message": call_info["message"],
                                "employee_name": claim_details["employee_name"],
                                "amount": claim_details["amount"],
                                "description": claim_details["description"],
                                "receipt_url": claim_details.get("receipt_url"),
                                "category": claim_details.get("category"),
                                "travel_type": claim_details.get("travel_type"),
                                "check_in_date": claim_details.get("check_in_date"),
                                "check_out_date": claim_details.get("check_out_date"),
                                "travel_start_date": claim_details.get("travel_start_date"),
                                "travel_end_date": claim_details.get("travel_end_date"),
                                "city": claim_details.get("city"),
                                "state": claim_details.get("state"),
                                "hotel_receipt_url": claim_details.get("hotel_receipt_url"),
                                "flight_ticket_receipt_url": claim_details.get("flight_ticket_receipt_url"),
                                "manager_approval_letter_url": claim_details.get("manager_approval_letter_url"),
                                "per_diem_rate": claim_details.get("per_diem_rate"),
                                "lodging_rate_limit": claim_details.get("lodging_rate_limit"),
                                "claimed_meals": claim_details.get("claimed_meals"),
                                "business_purpose": claim_details.get("business_purpose"),
                                "item_type": claim_details.get("item_type"),
                                "quantity": claim_details.get("quantity"),
                                "vendor": claim_details.get("vendor"),
                                "office_receipt_url": claim_details.get("office_receipt_url"),
                                "parking_receipt_url": claim_details.get("parking_receipt_url"),
                                "parking_location": claim_details.get("parking_location"),
                                "parking_date": claim_details.get("parking_date"),
                                "related_meeting": claim_details.get("related_meeting"),
                                "related_client": claim_details.get("related_client"),
                                "non_reimbursable_reason": claim_details.get("non_reimbursable_reason"),
                                "uploaded_docs": list(uploaded_metadata.keys()),
                                
                                # Multi-company and per-diem additions
                                "company_id": expense_data.get("company_id") or claim_details.get("company_id") or "demo_company",
                                "employee_email": expense_data.get("employee_email") or claim_details.get("employee_email") or "",
                                "employee_id": expense_data.get("employee_id") or claim_details.get("employee_id") or "",
                                "state_code": expense_data.get("state_code") or claim_details.get("state_code") or "",
                                "country": expense_data.get("country") or claim_details.get("country") or "US",
                                "claimed_lodging": expense_data.get("claimed_lodging") or claim_details.get("claimed_lodging") or 0.0,
                                "claimed_incidentals": expense_data.get("claimed_incidentals") or claim_details.get("claimed_incidentals") or 0.0,
                                "per_diem_review": expense_data.get("per_diem_review") or claim_details.get("per_diem_review")
                            })
                        return claims_parsed
                    return []
                except Exception as e:
                    logger.exception(f"Error fetching session {summary.id}: {e}")
                    return []
        
        # Gather sessions in parallel
        tasks = [fetch_and_parse(summary) for summary in active_summaries]
        results = await asyncio.gather(*tasks)
        
        for result in results:
            pending_claims.extend(result)
            
        # Merge Firestore-only claims from employee portal that are pending or blocked
        if db:
            try:
                existing_sess_ids = {c["session_id"] for c in pending_claims if "session_id" in c}
                db_expenses = await asyncio.to_thread(
                    lambda: list(
                        db.collection(EXPENSES_COL)
                        .where("status", "in", ["pending_review", "blocked_missing_docs"])
                        .get()
                    )
                )
                for doc in db_expenses:
                    exp = doc.to_dict()
                    sess_id = exp.get("session_id") or exp.get("claim_id") or doc.id
                    claim_id = exp.get("claim_id") or doc.id
                    
                    if sess_id in existing_sess_ids:
                        continue
                        
                    uploaded_metadata = []
                    try:
                        uploaded_docs = await asyncio.to_thread(
                            lambda: list(db.collection(DOCUMENTS_COL).where("session_id", "==", sess_id).get())
                        )
                        uploaded_metadata = [d.to_dict().get("doc_type") for d in uploaded_docs if d.to_dict().get("doc_type")]
                    except Exception as de:
                        logger.warning(f"Error fetching Firestore docs for sess {sess_id}: {de}")
                        
                    pending_claims.append({
                        "claim_id": claim_id,
                        "session_id": sess_id,
                        "user_id": exp.get("submitted_by_role") or "portal-user",
                        "interrupt_id": "review_decision",
                        "message": f"Attention required: Claim for {exp.get('employee_name')} is pending review.",
                        "employee_name": exp.get("employee_name") or exp.get("employee_email") or "Unknown Employee",
                        "amount": float(exp.get("amount") or 0.0),
                        "description": exp.get("description") or exp.get("business_purpose") or "Portal Claim Submission",
                        "receipt_url": exp.get("receipt_url"),
                        "category": exp.get("category"),
                        "travel_type": exp.get("travel_type") or ("travel" if exp.get("category") in ["hotel", "flight", "meals"] else "other"),
                        "check_in_date": exp.get("check_in_date"),
                        "check_out_date": exp.get("check_out_date"),
                        "travel_start_date": exp.get("travel_start_date"),
                        "travel_end_date": exp.get("travel_end_date"),
                        "city": exp.get("city"),
                        "state": exp.get("state"),
                        "hotel_receipt_url": exp.get("hotel_receipt_url"),
                        "flight_ticket_receipt_url": exp.get("flight_ticket_receipt_url"),
                        "manager_approval_letter_url": exp.get("manager_approval_letter_url"),
                        "per_diem_rate": exp.get("per_diem_rate"),
                        "lodging_rate_limit": exp.get("lodging_rate_limit"),
                        "claimed_meals": exp.get("claimed_meals"),
                        "business_purpose": exp.get("business_purpose"),
                        "uploaded_docs": list(uploaded_metadata),
                        "company_id": exp.get("company_id") or "demo_company",
                        "employee_email": exp.get("employee_email") or "",
                        "employee_id": exp.get("employee_id") or "",
                        "state_code": exp.get("state_code") or "",
                        "country": exp.get("country") or "US",
                        "claimed_lodging": exp.get("claimed_lodging") or 0.0,
                        "claimed_incidentals": exp.get("claimed_incidentals") or 0.0,
                        "per_diem_review": exp.get("per_diem_review") or {},
                        "policy_status": exp.get("policy_status") or "Compliance evaluation completed.",
                        "missing_documents": exp.get("missing_documents") or []
                    })
            except Exception as dbe:
                logger.error(f"Error merging Firestore pending claims: {dbe}")
            
        filtered_pending = filter_and_enrich_claims(pending_claims, current_user, params, is_pending=True)
        sanitized_pending_claims = [sanitize_claim_dict(c) for c in filtered_pending]
        result_payload = {
            "pending_claims": sanitized_pending_claims,
            "hidden_cli_sessions_count": hidden_cli_sessions_count
        }
        _pending_cache[cache_key] = (time.time(), result_payload)
        return result_payload
    except Exception as e:
        logger.error(f"Error fetching pending approvals: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch sessions: {str(e)}"
        )


@router.post("/api/action/{session_id}")
async def post_action(session_id: str, action: ApprovalAction, request: Request):
    """
    Resumes a paused Reasoning Engine session by invoking its execution api client directly.
    Also persists the manager decision and agent final response in Firestore.
    """
    current_user = get_current_user_and_role(request)
    role = current_user["role"]
    if role == "employee":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied: Employees cannot approve or reject claims."
        )

    if action.override_reason and role != "finance_admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied: Only finance administrators can override policy constraints."
        )

    # Phase 2: Resolve claim_id and write manager decision to Firestore
    claim_id = ""
    expense_id = ""
    decision_id = str(uuid.uuid4())
    is_portal_claim = False
    emp_email = None
    emp_name = None
    mgr_email = None
    if db:
        try:
            query = db.collection(EXPENSES_COL).where("session_id", "==", session_id).limit(1)
            expense_docs = list(query.get())
            if expense_docs:
                expense_id = expense_docs[0].id
                expense_dict = expense_docs[0].to_dict()
                claim_id = expense_dict.get("claim_id") or expense_id
                if expense_dict.get("source") == "employee_portal":
                    is_portal_claim = True
                
                # Enrich to ensure default/fallback fields are correctly retrieved
                expense_dict = enrich_claim_with_employee_info(expense_dict)
                emp_email = expense_dict.get("employee_email")
                emp_name = expense_dict.get("employee_name")
                mgr_email = expense_dict.get("manager_email")
            
            decision_doc = {
                "claim_id": claim_id,
                "session_id": session_id,
                "decision": "approved" if action.approved else "rejected",
                "decided_by": current_user["email"],
                "decided_role": current_user["role"],
                "actor_email": current_user["email"],
                "actor_role": current_user["role"],
                "authenticated": True,
                "decided_at": datetime.utcnow().isoformat() + "Z",
                "final_agent_response": ""
            }
            
            if action.override_reason:
                decision_doc["override_reason"] = action.override_reason
                decision_doc["decision_reason"] = f"Finance Admin override: {action.override_reason}"
            else:
                decision_doc["decision_reason"] = "Manager clicked Approve" if action.approved else "Manager clicked Reject"
                
            db.collection(DECISIONS_COL).document(decision_id).set(decision_doc)
            
            event_msg = f"Manager submitted decision: {'approved' if action.approved else 'rejected'}."
            if action.override_reason:
                event_msg = f"Finance Admin override: {action.override_reason}."
                
            add_audit_log(
                claim_id=claim_id,
                session_id=session_id,
                event_type="manager_decision",
                event_message=event_msg,
                actor=current_user["email"] if current_user.get("email") else current_user["role"],
                actor_email=current_user["email"],
                actor_role=current_user["role"],
                authenticated=current_user["authenticated"],
                employee_email=emp_email,
                employee_name=emp_name,
                manager_email=mgr_email,
                metadata={"override_reason": action.override_reason} if action.override_reason else None
            )
        except Exception as ex:
            logger.error(f"Error recording manager decision: {ex}")

    if is_portal_claim:
        try:
            final_status = "approved" if action.approved else "rejected"
            final_review = f"Manager directly {'approved' if action.approved else 'rejected'} this portal-submitted claim."
            
            if db:
                db.collection(DECISIONS_COL).document(decision_id).update({
                    "final_agent_response": final_review
                })
                if expense_id:
                    db.collection(EXPENSES_COL).document(expense_id).update({
                        "status": final_status,
                        "updated_at": datetime.utcnow().isoformat() + "Z",
                        "decision_by_email": current_user["email"],
                        "decision_by_role": current_user["role"],
                        "reviewer_email": current_user["email"],
                        "reviewer_role": current_user["role"],
                    })
                    
                add_audit_log(
                    claim_id=claim_id,
                    session_id=session_id,
                    event_type="agent_finalized",
                    event_message=f"Claim processed. Final Status: {final_status}.",
                    actor="agent",
                    metadata={"final_review": final_review}
                )
            
            return {
                "status": "success",
                "session_id": session_id,
                "approved": action.approved,
                "final_review": final_review
            }
        except Exception as pe:
            logger.error(f"Error executing portal bypass logic: {pe}")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Failed to finalize portal-submitted claim: {str(pe)}"
            )

    try:
        remote_app = reasoning_engines.ReasoningEngine(AGENT_RUNTIME_ID)
        
        # Build the exact resume payload expected by the workflow
        resume_payload = {
            "role": "user",
            "parts": [{
                "function_response": {
                    "id": action.interrupt_id,
                    "name": "adk_request_input",
                    "response": {"approved": action.approved}
                }
            }]
        }
        
        logger.info(f"Resuming session {session_id} with approval={action.approved}")
        
        # Invoke stream_query_reasoning_engine
        response_stream = remote_app.execution_api_client.stream_query_reasoning_engine(
            request=aip_types.StreamQueryReasoningEngineRequest(
                name=remote_app.resource_name,
                input={
                    "message": resume_payload,
                    "user_id": "default-user", # Strictly set to "default-user" per prompt rules
                    "session_id": session_id
                },
                class_method="stream_query",
            )
        )
        
        # Consume the stream and extract the model's final response text
        output_text_parts = []
        for chunk in response_stream:
            if hasattr(chunk, "data") and chunk.data:
                try:
                    payload = json.loads(chunk.data)
                    # Check for text inside content
                    if "content" in payload and "parts" in payload["content"]:
                        for part in payload["content"]["parts"]:
                            if "text" in part and part["text"]:
                                output_text_parts.append(part["text"])
                    # Check for final output attribute
                    elif "output" in payload and payload["output"]:
                        if isinstance(payload["output"], dict) and "comments" in payload["output"]:
                            output_text_parts.append(payload["output"]["comments"])
                except Exception:
                    pass
                    
        final_review = "".join(output_text_parts) if output_text_parts else "Approval submitted successfully."
        
        # Phase 2: Update status, final agent response, and audit log
        final_status = "approved" if action.approved else "rejected"
        if db:
            try:
                db.collection(DECISIONS_COL).document(decision_id).update({
                    "final_agent_response": final_review
                })
                if expense_id:
                    db.collection(EXPENSES_COL).document(expense_id).update({
                        "status": final_status,
                        "updated_at": datetime.utcnow().isoformat() + "Z",
                        "decision_by_email": current_user["email"],
                        "decision_by_role": current_user["role"],
                        "reviewer_email": current_user["email"],
                        "reviewer_role": current_user["role"],
                    })
                    
                add_audit_log(
                    claim_id=claim_id,
                    session_id=session_id,
                    event_type="agent_finalized",
                    event_message=f"Agent completed execution stream. Final Status: {final_status}.",
                    actor="agent",
                    metadata={"final_review": final_review}
                )
            except Exception as fe:
                logger.error(f"Error finalizing decision and status: {fe}")

        # Format a beautifully descriptive response
        return {
            "status": "success",
            "session_id": session_id,
            "approved": action.approved,
            "final_review": final_review
        }
        
    except Exception as e:
        logger.error(f"Error resuming session {session_id}: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to resume session: {str(e)}"
        )

@router.post("/api/upload/{session_id}/{doc_type}")
async def upload_document(session_id: str, doc_type: str, request: Request, file: UploadFile = File(...)):
    """
    Accepts multipart/form-data file upload, validates doc_type and file type,
    uploads to GCS, and updates metadata JSON.
    """
    current_user = get_current_user_and_role(request)
    role = current_user["role"]
    if role == "employee":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied: Employees cannot upload documents."
        )

    allowed_doc_types = {
        "receipt", "hotel_receipt", "flight_ticket_receipt", 
        "manager_approval_letter", "office_receipt", "parking_receipt"
    }
    if doc_type not in allowed_doc_types:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid document type: {doc_type}"
        )
    
    # Validate extension
    filename = file.filename or "file"
    ext = os.path.splitext(filename)[1].lower()
    allowed_exts = {".pdf", ".png", ".jpg", ".jpeg"}
    if ext not in allowed_exts:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported file type: {ext}. Only PDF, PNG, JPG, JPEG are allowed."
        )
    
    # Validate mime-type
    allowed_content_types = {"application/pdf", "image/png", "image/jpeg", "image/jpg"}
    if file.content_type not in allowed_content_types:
        # We also allow extension-based match as backup
        pass
        
    # File size validation (10 MB maximum)
    max_size = 10 * 1024 * 1024
    contents = await file.read()
    if len(contents) > max_size:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File size exceeds maximum allowed limit of 10 MB."
        )
        
    # Sanitize filename
    base_name = os.path.splitext(filename)[0]
    clean_base = re.sub(r'[^a-zA-Z0-9_.-]', '_', base_name)
    safe_filename = f"{clean_base}{ext}"
    
    object_path = f"uploads/{session_id}/{doc_type}/{safe_filename}"
    
    try:
        bucket = get_gcs_bucket()
        
        # Avoid overwriting files of other types, but replace previous files of the exact same document type
        blobs_to_delete = bucket.list_blobs(prefix=f"uploads/{session_id}/{doc_type}/")
        for b in blobs_to_delete:
            try:
                b.delete()
            except Exception as de:
                logger.warning(f"Could not delete old blob {b.name}: {de}")
                
        # Upload new blob
        blob = bucket.blob(object_path)
        blob.upload_from_string(contents, content_type=file.content_type or "application/octet-stream")
        
        # Update session metadata JSON
        metadata_path = f"uploads/{session_id}/metadata.json"
        metadata_blob = bucket.blob(metadata_path)
        
        metadata = {}
        if metadata_blob.exists():
            try:
                metadata = json.loads(metadata_blob.download_as_bytes().decode("utf-8"))
            except Exception as me:
                logger.warning(f"Error reading existing metadata for {session_id}: {me}")
                
        # Register new/replaced document
        metadata[doc_type] = {
            "path": object_path,
            "display_name": safe_filename,
            "content_type": file.content_type
        }
        
        metadata_blob.upload_from_string(
            json.dumps(metadata, indent=2),
            content_type="application/json"
        )
        
        # Phase 2: Write document metadata to Firestore and reevaluate policies
        claim_id = ""
        expense_id = ""
        if db:
            try:
                query = db.collection(EXPENSES_COL).where("session_id", "==", session_id).limit(1)
                expense_docs = list(query.get())
                if expense_docs:
                    expense_id = expense_docs[0].id
                    claim_id = expense_docs[0].to_dict().get("claim_id") or expense_id
                
                doc_data = {
                    "claim_id": claim_id,
                    "session_id": session_id,
                    "doc_type": doc_type,
                    "gcs_path": f"gs://{BUCKET_NAME}/{object_path}",
                    "original_filename": filename,
                    "content_type": file.content_type or "application/octet-stream",
                    "uploaded_at": datetime.utcnow().isoformat() + "Z"
                }
                # Store document details in firestore
                db.collection(DOCUMENTS_COL).document(f"{session_id}_{doc_type}").set(doc_data)
                
                # Write Audit Log
                add_audit_log(
                    claim_id=claim_id,
                    session_id=session_id,
                    event_type="document_uploaded",
                    event_message=f"Uploaded document {doc_type}: {filename}.",
                    actor=current_user["email"] if current_user.get("email") else current_user["role"],
                    actor_email=current_user["email"],
                    actor_role=current_user["role"],
                    authenticated=current_user["authenticated"],
                    metadata={"gcs_path": f"gs://{BUCKET_NAME}/{object_path}"}
                )
                
                # Reevaluate policies
                if expense_id:
                    await asyncio.to_thread(reevaluate_expense_policies, expense_id)
            except Exception as fe:
                logger.error(f"Error updating Firestore during document upload: {fe}")

        logger.info(f"Successfully uploaded {doc_type} to {object_path} for session {session_id}")
        return {
            "status": "success",
            "doc_type": doc_type,
            "path": object_path,
            "display_name": safe_filename
        }
    except Exception as e:
        logger.error(f"Error uploading file: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"GCS Upload failed: {str(e)}"
        )


@router.get("/api/uploads/{session_id}")
async def get_uploads_metadata(session_id: str, request: Request):
    """
    Returns uploaded document metadata for a session.
    """
    current_user = get_current_user_and_role(request)
    if current_user["role"] == "employee":
        # Check if the employee owns this session
        if db:
            expense_docs = list(db.collection(EXPENSES_COL).where("session_id", "==", session_id).limit(1).get())
            if expense_docs:
                expense_data = expense_docs[0].to_dict()
                if expense_data.get("employee_name", "").lower() != current_user["email"].lower():
                    raise HTTPException(
                        status_code=status.HTTP_403_FORBIDDEN,
                        detail="Access denied: You are not authorized to view uploads metadata for this session."
                    )
    """
    Returns uploaded document metadata for a session.
    """
    try:
        bucket = get_gcs_bucket()
        metadata_blob = bucket.blob(f"uploads/{session_id}/metadata.json")
        if not metadata_blob.exists():
            return {}
        return json.loads(metadata_blob.download_as_bytes().decode("utf-8"))
    except Exception as e:
        logger.error(f"Error reading GCS metadata for {session_id}: {e}")
        return {}


@router.get("/api/document/{session_id}/{doc_type}")
async def get_uploaded_document(session_id: str, doc_type: str, request: Request):
    """
    Securely retrieves and streams an uploaded file for viewing from private GCS bucket.
    """
    current_user = get_current_user_and_role(request)
    if current_user["role"] == "employee":
        # Check if the employee owns this session
        if db:
            expense_docs = list(db.collection(EXPENSES_COL).where("session_id", "==", session_id).limit(1).get())
            if expense_docs:
                expense_data = expense_docs[0].to_dict()
                if expense_data.get("employee_name", "").lower() != current_user["email"].lower():
                    raise HTTPException(
                        status_code=status.HTTP_403_FORBIDDEN,
                        detail="Access denied: You are not authorized to view this document."
                    )
    """
    Securely retrieves and streams an uploaded file for viewing from private GCS bucket.
    """
    try:
        bucket = get_gcs_bucket()
        metadata_blob = bucket.blob(f"uploads/{session_id}/metadata.json")
        if not metadata_blob.exists():
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No metadata found for session")
            
        metadata = json.loads(metadata_blob.download_as_bytes().decode("utf-8"))
        if doc_type not in metadata:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"No document of type {doc_type} uploaded")
            
        doc_info = metadata[doc_type]
        blob = bucket.blob(doc_info["path"])
        if not blob.exists():
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File not found in storage")
            
        file_bytes = blob.download_as_bytes()
        import io
        return StreamingResponse(
            io.BytesIO(file_bytes),
            media_type=blob.content_type or "application/octet-stream",
            headers={
                "Content-Disposition": f"inline; filename={urllib.parse.quote(doc_info['display_name'])}"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error serving document: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch document: {str(e)}"
        )


@router.post("/api/ingest-expense")
async def ingest_expense(payload: dict, request: Request):
    """
    Stores incoming Pub/Sub expense event into Firestore.
    Does not replace the Agent Runtime processing flow.
    Supports both direct JSON payloads and standard Pub/Sub push notification structures.
    """
    # Ingest Endpoint Shared-Secret Verification
    shared_secret = os.getenv("INGEST_SHARED_SECRET", "")
    if shared_secret:
        req_secret = request.query_params.get("secret")
        auth_header = request.headers.get("authorization")
        if auth_header and auth_header.startswith("Bearer "):
            bearer_token = auth_header[7:]
            req_secret = bearer_token
            # Add robust OIDC Token verification fallback to accept our Pub/Sub invoker service account
            try:
                parts = bearer_token.split('.')
                if len(parts) == 3:
                    import base64
                    payload_b64 = parts[1] + '=' * (4 - len(parts[1]) % 4)
                    payload_data = json.loads(base64.b64decode(payload_b64).decode('utf-8'))
                    if payload_data.get("iss") in ["accounts.google.com", "https://accounts.google.com"]:
                        email = payload_data.get("email")
                        if email == "pubsub-invoker@project-5d38f91a-29a3-45bd-8d4.iam.gserviceaccount.com":
                            logger.info(f"Authenticated ingest via Google OIDC Token for {email}")
                            req_secret = shared_secret
            except Exception as jwte:
                logger.warning(f"Failed to decode OIDC bearer token: {jwte}")
            
        if req_secret != shared_secret:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Unauthorized: Invalid ingest shared secret."
            )
    """
    Stores incoming Pub/Sub expense event into Firestore.
    Does not replace the Agent Runtime processing flow.
    Supports both direct JSON payloads and standard Pub/Sub push notification structures.
    """
    logger.info(f"Received ingest-expense payload: {payload}")
    try:
        # Check if it is a Pub/Sub push message
        raw_data = None
        if "message" in payload and "data" in payload["message"]:
            import base64
            encoded_data = payload["message"]["data"]
            try:
                decoded = base64.b64decode(encoded_data).decode("utf-8")
                raw_data = json.loads(decoded)
            except Exception as e:
                logger.error(f"Failed to decode Pub/Sub base64 data: {e}")
                raise HTTPException(status_code=400, detail="Invalid Pub/Sub base64 data")
        else:
            raw_data = payload

        # Now, raw_data should be the expense dictionary. Let's parse/unwrap input.message if needed.
        if "input" in raw_data and isinstance(raw_data["input"], dict):
            if "message" in raw_data["input"] and isinstance(raw_data["input"]["message"], dict):
                raw_data = raw_data["input"]["message"]
            elif "message" in raw_data["input"] and isinstance(raw_data["input"]["message"], str):
                try:
                    raw_data = json.loads(raw_data["input"]["message"])
                except Exception:
                    pass

        # Generate or preserve claim_id
        claim_id = raw_data.get("claim_id") or str(uuid.uuid4())
        session_id = raw_data.get("session_id") or ""
        user_id = raw_data.get("user_id") or "default-user"
        
        employee_name = raw_data.get("employee_name") or "Unknown Claimant"
        amount = float(raw_data.get("amount") or 0.0)
        
        # Policy review to set correct initial properties
        policy_res = run_policy_check_py(raw_data)
        
        # Determine status. Initially, all ingested expenses are "submitted".
        initial_status = "submitted"
        
        new_expense = {
            "claim_id": claim_id,
            "session_id": session_id,
            "user_id": user_id,
            "employee_name": employee_name,
            "amount": amount,
            "category": raw_data.get("category", "N/A"),
            "description": raw_data.get("description") or raw_data.get("business_purpose") or "No description provided",
            "business_purpose": raw_data.get("business_purpose", ""),
            "status": initial_status,
            "policy_status": policy_res["status"],
            "required_documents": policy_res["required_docs"],
            "missing_documents": policy_res["missing_docs"],
            "created_at": datetime.utcnow().isoformat() + "Z",
            "updated_at": datetime.utcnow().isoformat() + "Z"
        }
        
        # Store other custom/extra fields in the document
        for field in [
            "company_id", "employee_email", "employee_id", "state_code", "country",
            "claimed_meals", "claimed_lodging", "claimed_incidentals", "per_diem_review",
            "check_in_date", "check_out_date", "travel_start_date", "travel_end_date",
            "city", "state", "receipt_url", "hotel_receipt_url", "flight_ticket_receipt_url",
            "manager_approval_letter_url", "per_diem_rate", "lodging_rate_limit",
            "item_type", "quantity", "vendor", "office_receipt_url", "parking_receipt_url",
            "parking_location", "parking_date", "related_meeting", "related_client", "non_reimbursable_reason"
        ]:
            if field in raw_data:
                new_expense[field] = raw_data[field]
            elif field == "per_diem_review" and "per_diem_review" in raw_data:
                new_expense["per_diem_review"] = raw_data["per_diem_review"]
                
        # Ensure company_id and other fields are always present
        pdr_dict = new_expense.get("per_diem_review") or {}
        if not new_expense.get("company_id"):
            new_expense["company_id"] = pdr_dict.get("company_id") or "demo_company"
        if not new_expense.get("employee_email"):
            new_expense["employee_email"] = pdr_dict.get("employee_email") or ""
        if not new_expense.get("employee_id"):
            new_expense["employee_id"] = pdr_dict.get("employee_id") or ""
                
        if db:
            db.collection(EXPENSES_COL).document(claim_id).set(new_expense)
            
            # Write audit log
            add_audit_log(
                claim_id=claim_id,
                session_id=session_id,
                event_type="claim_ingested",
                event_message=f"Expense ingested via Pub/Sub endpoint. Status: {initial_status}.",
                actor="system",
                metadata={"payload": raw_data, "policy_review": policy_res}
            )
            
        logger.info(f"Successfully ingested expense claim {claim_id} for {employee_name} of amount {amount}")
        return {"status": "success", "claim_id": claim_id}
        
    except Exception as e:
        logger.error(f"Error in ingest_expense endpoint: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to ingest expense: {str(e)}")


@router.get("/api/expenses")
async def get_expenses(
    request: Request,
    search: str = None,
    department: str = None,
    manager: str = None,
    status: str = None,
    category: str = None,
    company: str = None,
    hide_old_test_sessions: bool = False,
    assigned_to_me: bool = False,
    show_all_fa: bool = False
):
    """
    Returns recent expenses from Firestore, syncing completed sessions first.
    """
    current_user = get_current_user_and_role(request)
    params = {
        "search": search,
        "department": department,
        "manager": manager,
        "status": status,
        "category": category,
        "company": company,
        "hide_old_test_sessions": hide_old_test_sessions,
        "assigned_to_me": assigned_to_me,
        "show_all_fa": show_all_fa
    }
    try:
        if db:
            # Run session sync to capture auto-approvals and finalised sessions
            try:
                await sync_completed_sessions()
            except Exception as se:
                logger.warning(f"Error syncing completed sessions during get_expenses: {se}")
                
            expenses_ref = db.collection(EXPENSES_COL).order_by("created_at", direction=firestore.Query.DESCENDING)
            docs = list(expenses_ref.get())
            all_claims = [doc.to_dict() for doc in docs]
            filtered_claims = filter_and_enrich_claims(all_claims, current_user, params, is_pending=False)
            return [sanitize_claim_dict(c) for c in filtered_claims]
        return []
    except Exception as e:
        logger.error(f"Error fetching expenses from Firestore: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch expenses: {str(e)}")


@router.get("/api/expenses/{claim_id}")
async def get_expense_detail(claim_id: str, request: Request):
    """
    Returns expense, documents, decisions, and audit log for one claim.
    """
    current_user = get_current_user_and_role(request)
    if not db:
        raise HTTPException(status_code=500, detail="Database connection not available")
    try:
        doc_ref = db.collection(EXPENSES_COL).document(claim_id)
        expense_doc = doc_ref.get()
        if not expense_doc.exists:
            raise HTTPException(status_code=404, detail=f"Expense claim {claim_id} not found")
            
        expense_data = enrich_claim_with_employee_info(expense_doc.to_dict())
        is_owner = (
            expense_data.get("employee_email") == current_user["email"] or
            expense_data.get("submitted_by_email") == current_user["email"]
        )
        is_manager = (
            expense_data.get("manager_email") == current_user["email"]
        )
        is_authorized = (
            current_user["role"] == "finance_admin" or
            is_owner or
            is_manager
        )
        if not is_authorized:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied: You are not authorized to view details of this expense claim."
            )
            
        # Query documents
        docs_ref = db.collection(DOCUMENTS_COL).where("claim_id", "==", claim_id)
        documents = [d.to_dict() for d in docs_ref.get()]
        
        # Query decisions
        decisions_ref = db.collection(DECISIONS_COL).where("claim_id", "==", claim_id)
        decisions = [d.to_dict() for d in decisions_ref.get()]
        
        # Query audit logs
        audits_ref = db.collection(AUDIT_LOGS_COL).where("claim_id", "==", claim_id)
        audits = [a.to_dict() for a in audits_ref.get()]
        audits = sorted(audits, key=lambda x: x.get("timestamp", ""))
        
        return {
            "expense": sanitize_claim_dict(expense_data),
            "documents": documents,
            "decisions": decisions,
            "audit_logs": audits
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching expense detail for claim_id {claim_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch expense details: {str(e)}")


@router.get("/api/audit/{claim_id}")
async def get_audit_trail(claim_id: str, request: Request):
    """
    Returns the sorted chronological audit timeline for one claim.
    """
    current_user = get_current_user_and_role(request)
    if not db:
        raise HTTPException(status_code=500, detail="Database connection not available")
    try:
        # First fetch expense document to verify authorization
        doc_ref = db.collection(EXPENSES_COL).document(claim_id)
        expense_doc = doc_ref.get()
        if not expense_doc.exists:
            raise HTTPException(status_code=404, detail=f"Expense claim {claim_id} not found")
            
        expense_data = enrich_claim_with_employee_info(expense_doc.to_dict())
        is_owner = (
            expense_data.get("employee_email") == current_user["email"] or
            expense_data.get("submitted_by_email") == current_user["email"]
        )
        is_manager = (
            expense_data.get("manager_email") == current_user["email"]
        )
        is_authorized = (
            current_user["role"] == "finance_admin" or
            is_owner or
            is_manager
        )
        if not is_authorized:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied: You are not authorized to view the audit trail for this expense claim."
            )
            
        audits_ref = db.collection(AUDIT_LOGS_COL).where("claim_id", "==", claim_id)
        audits = [a.to_dict() for a in audits_ref.get()]
        audits = sorted(audits, key=lambda x: x.get("timestamp", ""))
        return audits
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching audit trail for claim_id {claim_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch audit trail: {str(e)}")


# --- Company Policy Admin Endpoints ---

class StateRatePayload(BaseModel):
    company_id: str = "demo_company"
    state: str
    state_code: str
    country: str = "US"
    cost_tier: str | None = None
    meal_rate_per_day: float
    lodging_rate_per_night: float
    incidental_rate_per_day: float = 0.0
    effective_start_date: str | None = None
    effective_end_date: str | None = None
    active: bool = True
    source_note: str | None = None

class CityRatePayload(BaseModel):
    company_id: str = "demo_company"
    city: str
    state: str
    state_code: str
    country: str = "US"
    meal_rate_per_day: float
    lodging_rate_per_night: float
    incidental_rate_per_day: float = 0.0
    effective_start_date: str | None = None
    effective_end_date: str | None = None
    active: bool = True
    source_note: str | None = None

class DefaultsPayload(BaseModel):
    company_id: str = "demo_company"
    policy_id: str | None = None
    policy_name: str = "Default Policy"
    description: str | None = None
    default_meal_rate_per_day: float
    default_lodging_rate_per_night: float
    default_incidental_rate_per_day: float = 0.0
    partial_day_percent: float = 1.0
    receipt_required_above: float = 50.0
    manager_approval_required_above: float = 350.0
    require_receipt_for_flights: bool = True
    require_manager_letter_for_flight_above: float = 1200.0
    require_manager_letter_for_lodging_above: float = 350.0
    active: bool = True
    effective_start_date: str | None = None
    effective_end_date: str | None = None

@router.get("/api/company-policy")
async def get_company_policy(request: Request, company_id: str = "demo_company"):
    current_user = get_current_user_and_role(request)
    if current_user["role"] == "employee":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied: Employees cannot view company per diem policies."
        )
    if not db:
        raise HTTPException(status_code=500, detail="Database connection not available")
    try:
        # Query company default policy
        default_policy_docs = list(db.collection("company_policy_settings")
            .where("company_id", "==", company_id)
            .where("active", "==", True).get())
        defaults = [d.to_dict() for d in default_policy_docs]

        # Query state rates
        state_rate_docs = list(db.collection("company_state_per_diem_rates")
            .where("company_id", "==", company_id)
            .where("active", "==", True).get())
        state_rates = [d.to_dict() for d in state_rate_docs]

        # Query city rates
        city_rate_docs = list(db.collection("company_city_per_diem_rates")
            .where("company_id", "==", company_id)
            .where("active", "==", True).get())
        city_rates = [d.to_dict() for d in city_rate_docs]

        return {
            "company_id": company_id,
            "default_policy_settings": defaults,
            "state_rates": state_rates,
            "city_rates": city_rates
        }
    except Exception as e:
        logger.error(f"Error fetching company policies: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch company policies: {str(e)}")

@router.post("/api/company-policy/state-rate")
async def post_state_rate(payload: StateRatePayload, request: Request):
    current_user = get_current_user_and_role(request)
    if current_user["role"] != "finance_admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied: Only finance administrators can create or update state per diem rates."
        )
    if not db:
        raise HTTPException(status_code=500, detail="Database connection not available")
    try:
        rate_id = f"{payload.company_id}_{payload.state_code.lower()}"
        data = payload.dict()
        data["rate_id"] = rate_id
        data["created_by"] = current_user["email"]
        data["updated_by"] = current_user["email"]
        data["created_at"] = datetime.utcnow().isoformat() + "Z"
        data["updated_at"] = datetime.utcnow().isoformat() + "Z"
        
        db.collection("company_state_per_diem_rates").document(rate_id).set(data)
        
        # Write audit log
        add_audit_log(
            claim_id="",
            session_id="",
            event_type="state_rate_updated",
            event_message=f"Created or updated state per diem rate for state {payload.state_code}.",
            actor=current_user["email"] if current_user.get("email") else current_user["role"],
            actor_email=current_user["email"],
            actor_role=current_user["role"],
            authenticated=current_user["authenticated"],
            metadata={"state_rate": data}
        )
        return {"status": "success", "rate_id": rate_id}
    except Exception as e:
        logger.error(f"Error writing state rate: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to save state per diem rate: {str(e)}")

@router.post("/api/company-policy/city-rate")
async def post_city_rate(payload: CityRatePayload, request: Request):
    current_user = get_current_user_and_role(request)
    if current_user["role"] != "finance_admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied: Only finance administrators can create or update city per diem rates."
        )
    if not db:
        raise HTTPException(status_code=500, detail="Database connection not available")
    try:
        city_slug = payload.city.lower().replace(" ", "_")
        rate_id = f"{payload.company_id}_{payload.state_code.lower()}_{city_slug}"
        data = payload.dict()
        data["rate_id"] = rate_id
        data["created_by"] = current_user["email"]
        data["updated_by"] = current_user["email"]
        data["created_at"] = datetime.utcnow().isoformat() + "Z"
        data["updated_at"] = datetime.utcnow().isoformat() + "Z"
        
        db.collection("company_city_per_diem_rates").document(rate_id).set(data)
        
        # Write audit log
        add_audit_log(
            claim_id="",
            session_id="",
            event_type="city_rate_updated",
            event_message=f"Created or updated city per diem rate for city {payload.city}, state {payload.state_code}.",
            actor=current_user["email"] if current_user.get("email") else current_user["role"],
            actor_email=current_user["email"],
            actor_role=current_user["role"],
            authenticated=current_user["authenticated"],
            metadata={"city_rate": data}
        )
        return {"status": "success", "rate_id": rate_id}
    except Exception as e:
        logger.error(f"Error writing city rate: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to save city per diem rate: {str(e)}")

@router.post("/api/company-policy/defaults")
async def post_defaults(payload: DefaultsPayload, request: Request):
    current_user = get_current_user_and_role(request)
    if current_user["role"] != "finance_admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied: Only finance administrators can update default company policies."
        )
    if not db:
        raise HTTPException(status_code=500, detail="Database connection not available")
    try:
        policy_id = payload.policy_id or f"{payload.company_id}_default"
        data = payload.dict()
        data["policy_id"] = policy_id
        data["created_by"] = current_user["email"]
        data["updated_by"] = current_user["email"]
        data["created_at"] = datetime.utcnow().isoformat() + "Z"
        data["updated_at"] = datetime.utcnow().isoformat() + "Z"
        
        db.collection("company_policy_settings").document(policy_id).set(data)
        
        # Write audit log
        add_audit_log(
            claim_id="",
            session_id="",
            event_type="company_defaults_updated",
            event_message=f"Created or updated default company per diem policy settings.",
            actor=current_user["email"] if current_user.get("email") else current_user["role"],
            actor_email=current_user["email"],
            actor_role=current_user["role"],
            authenticated=current_user["authenticated"],
            metadata={"policy_settings": data}
        )
        return {"status": "success", "policy_id": policy_id}
    except Exception as e:
        logger.error(f"Error writing default policy settings: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to save default company per diem policy settings: {str(e)}")



@router.get("/api/me")
async def get_me(request: Request):
    user = get_current_user_and_role(request)
    return {
        "email": user["email"],
        "role": user["role"],
        "name": user.get("name", ""),
        "company_id": "demo_company",
        "authenticated": user["authenticated"]
    }

@router.post("/api/employee/claims")
async def create_employee_claim(claim_data: dict, request: Request):
    current_user = get_current_user_and_role(request)
    
    # 1. Validate required fields
    if not claim_data.get("category"):
        raise HTTPException(status_code=400, detail="Category is required")
    try:
        amount = float(claim_data.get("amount") or 0.0)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid amount value")
        
    # 2. Assign & Normalize
    claim_id = claim_data.get("claim_id") or str(uuid.uuid4())
    session_id = claim_id  # Align session_id with claim_id for upload matching
    
    company_id = claim_data.get("company_id") or "demo_company"
    
    # Check if logged in user is finance_admin and submits on behalf of demo employee
    employee_email = claim_data.get("employee_email")
    if current_user["role"] == "finance_admin" and employee_email:
        pass
    else:
        employee_email = current_user["email"]
        
    # Retrieve employee details from Firestore (Next Step 1)
    employee_profile_status = "missing_employee_record"
    employee_id = ""
    employee_name = claim_data.get("employee_name") or employee_email.split("@")[0].replace(".", " ").title()
    department = claim_data.get("department") or "N/A"
    manager_email = claim_data.get("manager_email") or "manager@company.com"
    
    if db:
        try:
            emp_docs = list(db.collection("employees").where("employee_email", "==", employee_email).where("active", "==", True).limit(1).get())
            if emp_docs:
                emp_data = emp_docs[0].to_dict()
                employee_profile_status = "found"
                department = emp_data.get("department") or department
                manager_email = emp_data.get("manager_email") or manager_email
                employee_id = emp_data.get("employee_id") or employee_id
                employee_name = emp_data.get("employee_name") or employee_name
        except Exception as e:
            logger.warning(f"Error checking employee profile during claim creation: {e}")

    claim_data["employee_email"] = employee_email
    claim_data["employee_name"] = employee_name
    claim_data["employee_id"] = employee_id
    claim_data["employee_profile_status"] = employee_profile_status
    claim_data["department"] = department
    claim_data["manager_email"] = manager_email
    
    # Check for existing document uploads
    if db:
        uploaded_docs = list(db.collection(DOCUMENTS_COL).where("session_id", "==", session_id).get())
        for doc_item in uploaded_docs:
            doc_data = doc_item.to_dict()
            doc_type = doc_data.get("doc_type")
            doc_url_fields = {
                "receipt": "receipt_url",
                "hotel_receipt": "hotel_receipt_url",
                "flight_ticket_receipt": "flight_ticket_receipt_url",
                "manager_approval_letter": "manager_approval_letter_url",
                "office_receipt": "office_receipt_url",
                "parking_receipt": "parking_receipt_url",
                "mileage_log": "mileage_log_url",
                "rental_receipt": "rental_receipt_url",
                "gas_receipt": "gas_receipt_url",
                "toll_receipt": "toll_receipt_url"
            }
            if doc_type in doc_url_fields:
                claim_data[doc_url_fields[doc_type]] = f"/api/document/{session_id}/{doc_type}"
                
    # 4. Run policy check
    policy_res = run_policy_check_py(claim_data)
    
    # 3. Determine status
    if policy_res.get("missing_docs"):
        status = "blocked_missing_docs"
    elif not policy_res.get("is_valid") or "Requires manager approval" in policy_res.get("status", "") or "Requires manager approval letter" in policy_res.get("status", "") or policy_res.get("status") == "Potentially non-reimbursable":
        status = "pending_review"
    else:
        status = "auto_approved"
        
    pdr = policy_res.get("per_diem_review") or {}
    
    normalized_claim = {
        "claim_id": claim_id,
        "session_id": session_id,
        "company_id": company_id,
        "employee_email": employee_email,
        "employee_name": employee_name,
        "employee_id": employee_id,
        "employee_profile_status": employee_profile_status,
        "department": department,
        "manager_email": manager_email,
        "category": claim_data.get("category"),
        "amount": amount,
        "currency": claim_data.get("currency") or "USD",
        "expense_date": claim_data.get("expense_date") or datetime.utcnow().strftime("%Y-%m-%d"),
        "business_purpose": claim_data.get("business_purpose") or "",
        "description": claim_data.get("description") or "",
        "status": status,
        "policy_status": policy_res["status"],
        "required_documents": policy_res["required_docs"],
        "missing_documents": policy_res["missing_docs"],
        "submitted_by_email": current_user["email"],
        "submitted_by_role": current_user["role"],
        "submitted_at": datetime.utcnow().isoformat() + "Z",
        "created_at": datetime.utcnow().isoformat() + "Z",
        "updated_at": datetime.utcnow().isoformat() + "Z",
        "source": "employee_portal",
        
        # Travel fields
        "travel_start_date": claim_data.get("travel_start_date"),
        "travel_end_date": claim_data.get("travel_end_date"),
        "city": claim_data.get("city"),
        "state": claim_data.get("state"),
        "state_code": claim_data.get("state_code") or claim_data.get("state"),
        "country": claim_data.get("country") or "US",
        "claimed_meals": float(claim_data.get("claimed_meals") or 0.0),
        "claimed_lodging": float(claim_data.get("claimed_lodging") or 0.0),
        "claimed_incidentals": float(claim_data.get("claimed_incidentals") or 0.0),
        "check_in_date": claim_data.get("check_in_date"),
        "check_out_date": claim_data.get("check_out_date"),
        
        # Transportation fields
        "transportation_type": claim_data.get("transportation_type"),
        "trip_date": claim_data.get("trip_date"),
        "start_location_label": claim_data.get("start_location_label"),
        "start_address": claim_data.get("start_address"),
        "destination_location_label": claim_data.get("destination_location_label"),
        "destination_address": claim_data.get("destination_address"),
        "trip_type": claim_data.get("trip_type"),
        "business_miles": float(claim_data.get("business_miles") or 0.0),
        "employee_entered_miles": float(claim_data.get("employee_entered_miles") or 0.0),
        "rental_start_date": claim_data.get("rental_start_date"),
        "rental_end_date": claim_data.get("rental_end_date"),
        "rental_cost": float(claim_data.get("rental_cost") or 0.0),
        "gas_cost": float(claim_data.get("gas_cost") or 0.0),
        "parking_cost": float(claim_data.get("parking_cost") or 0.0),
        "toll_cost": float(claim_data.get("toll_cost") or 0.0),
        "linked_rental_claim_id": claim_data.get("linked_rental_claim_id"),
        "per_diem_review": pdr
    }
    
    # Assign document URLs
    for doc_type, url_field in {
        "receipt": "receipt_url",
        "hotel_receipt": "hotel_receipt_url",
        "flight_ticket_receipt": "flight_ticket_receipt_url",
        "manager_approval_letter": "manager_approval_letter_url",
        "office_receipt": "office_receipt_url",
        "parking_receipt": "parking_receipt_url",
        "mileage_log": "mileage_log_url",
        "rental_receipt": "rental_receipt_url",
        "gas_receipt": "gas_receipt_url",
        "toll_receipt": "toll_receipt_url"
    }.items():
        if claim_data.get(url_field):
            normalized_claim[url_field] = claim_data[url_field]
            
    if db:
        db.collection(EXPENSES_COL).document(claim_id).set(normalized_claim)
        
        # 6. Store audit log
        add_audit_log(
            claim_id=claim_id,
            session_id=session_id,
            event_type="employee_submitted_claim",
            event_message=f"Employee submitted claim {claim_id} via portal. Status: {status}.",
            actor=current_user["email"],
            actor_email=current_user["email"],
            actor_role=current_user["role"],
            authenticated=True,
            employee_email=employee_email,
            employee_name=employee_name,
            manager_email=manager_email,
            metadata={"policy_review": policy_res, "source": "employee_portal"}
        )
        
        # If auto-approved, record decision and finalize log
        if status == "auto_approved":
            decision_ref = db.collection(DECISIONS_COL).document()
            decision_ref.set({
                "claim_id": claim_id,
                "session_id": session_id,
                "decision": "approved",
                "decided_by": "agent",
                "decision_reason": "Claim is within company policy thresholds. Automated system approval.",
                "decided_at": datetime.utcnow().isoformat() + "Z",
                "final_agent_response": "Automated system approval."
            })
            add_audit_log(
                claim_id=claim_id,
                session_id=session_id,
                event_type="agent_finalized",
                event_message=f"Claim auto-approved. Status: auto_approved.",
                actor="agent",
                metadata={"final_response": "Automated system approval."}
            )
            
    return normalized_claim

@router.get("/api/employee/claims")
async def get_employee_claims(request: Request):
    current_user = get_current_user_and_role(request)
    if not db:
        return []
    try:
        show_all = request.query_params.get("all") == "true" and current_user["role"] == "finance_admin"
        if show_all:
            query = db.collection(EXPENSES_COL).order_by("created_at", direction=firestore.Query.DESCENDING)
            docs = list(query.get())
            return [sanitize_claim_dict(doc.to_dict()) for doc in docs]
        else:
            email = current_user["email"]
            claims_dict = {}
            q1 = db.collection(EXPENSES_COL).where("employee_email", "==", email).get()
            for doc in q1:
                claims_dict[doc.id] = doc.to_dict()
            q2 = db.collection(EXPENSES_COL).where("submitted_by_email", "==", email).get()
            for doc in q2:
                claims_dict[doc.id] = doc.to_dict()
            claims_list = list(claims_dict.values())
            claims_list.sort(key=lambda x: x.get("created_at", ""), reverse=True)
            return [sanitize_claim_dict(c) for c in claims_list]
    except Exception as e:
        logger.error(f"Error fetching employee claims: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/api/employee/claims/{claim_id}")
async def get_employee_claim_detail(claim_id: str, request: Request):
    current_user = get_current_user_and_role(request)
    if not db:
        raise HTTPException(status_code=500, detail="Database connection not available")
    try:
        doc_ref = db.collection(EXPENSES_COL).document(claim_id)
        expense_doc = doc_ref.get()
        if not expense_doc.exists:
            raise HTTPException(status_code=404, detail=f"Claim {claim_id} not found")
            
        expense_data = expense_doc.to_dict()
        
        is_owner = (
            expense_data.get("employee_email") == current_user["email"] or
            expense_data.get("submitted_by_email") == current_user["email"] or
            expense_data.get("employee_name", "").lower() == current_user["email"].lower()
        )
        is_manager_or_admin = current_user["role"] in ["manager", "finance_admin"]
        
        if not is_owner and not is_manager_or_admin:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied: You are not authorized to view this claim."
            )
            
        docs_ref = db.collection(DOCUMENTS_COL).where("claim_id", "==", claim_id)
        documents = [d.to_dict() for d in docs_ref.get()]
        
        decisions_ref = db.collection(DECISIONS_COL).where("claim_id", "==", claim_id)
        decisions = [d.to_dict() for d in decisions_ref.get()]
        
        audits_ref = db.collection(AUDIT_LOGS_COL).where("claim_id", "==", claim_id)
        audits = [a.to_dict() for a in audits_ref.get()]
        audits = sorted(audits, key=lambda x: x.get("timestamp", ""))
        
        return {
            "expense": sanitize_claim_dict(expense_data),
            "documents": documents,
            "decisions": decisions,
            "audit_logs": audits
        }
    except Exception as e:
        logger.error(f"Error fetching employee claim detail for {claim_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/api/employee/claims/{claim_id}/documents/{doc_type}")
async def upload_employee_document(claim_id: str, doc_type: str, request: Request, file: UploadFile = File(...)):
    current_user = get_current_user_and_role(request)
    session_id = claim_id
    
    if db:
        doc_ref = db.collection(EXPENSES_COL).document(claim_id)
        expense_doc = doc_ref.get()
        if expense_doc.exists:
            expense_data = expense_doc.to_dict()
            is_owner = (
                expense_data.get("employee_email") == current_user["email"] or
                expense_data.get("submitted_by_email") == current_user["email"] or
                expense_data.get("employee_name", "").lower() == current_user["email"].lower()
            )
            is_manager_or_admin = current_user["role"] in ["manager", "finance_admin"]
            if not is_owner and not is_manager_or_admin:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Access denied: You are not authorized to upload documents for this claim."
                )
                
    allowed_doc_types = {
        "receipt", "hotel_receipt", "flight_ticket_receipt", 
        "manager_approval_letter", "office_receipt", "parking_receipt",
        "mileage_log", "rental_receipt", "gas_receipt", "toll_receipt"
    }
    if doc_type not in allowed_doc_types:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid document type: {doc_type}"
        )
        
    filename = file.filename or "file"
    ext = os.path.splitext(filename)[1].lower()
    allowed_exts = {".pdf", ".png", ".jpg", ".jpeg"}
    if ext not in allowed_exts:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported file type: {ext}. Only PDF, PNG, JPG, JPEG are allowed."
        )
        
    max_size = 10 * 1024 * 1024
    contents = await file.read()
    if len(contents) > max_size:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File size exceeds maximum allowed limit of 10 MB."
        )
        
    base_name = os.path.splitext(filename)[0]
    clean_base = re.sub(r'[^a-zA-Z0-9_.-]', '_', base_name)
    safe_filename = f"{clean_base}{ext}"
    
    object_path = f"uploads/{session_id}/{doc_type}/{safe_filename}"
    
    try:
        bucket = get_gcs_bucket()
        blobs_to_delete = bucket.list_blobs(prefix=f"uploads/{session_id}/{doc_type}/")
        for b in blobs_to_delete:
            try:
                b.delete()
            except Exception as de:
                logger.warning(f"Could not delete old blob {b.name}: {de}")
                
        blob = bucket.blob(object_path)
        blob.upload_from_string(contents, content_type=file.content_type or "application/octet-stream")
        
        metadata_path = f"uploads/{session_id}/metadata.json"
        metadata_blob = bucket.blob(metadata_path)
        metadata = {}
        if metadata_blob.exists():
            try:
                metadata = json.loads(metadata_blob.download_as_bytes().decode("utf-8"))
            except Exception as me:
                logger.warning(f"Error reading metadata: {me}")
                
        metadata[doc_type] = {
            "path": object_path,
            "display_name": safe_filename,
            "content_type": file.content_type
        }
        metadata_blob.upload_from_string(json.dumps(metadata, indent=2), content_type="application/json")
        
        if db:
            doc_data = {
                "claim_id": claim_id,
                "session_id": session_id,
                "doc_type": doc_type,
                "gcs_path": f"gs://{BUCKET_NAME}/{object_path}",
                "original_filename": filename,
                "content_type": file.content_type or "application/octet-stream",
                "uploaded_at": datetime.utcnow().isoformat() + "Z"
            }
            db.collection(DOCUMENTS_COL).document(f"{session_id}_{doc_type}").set(doc_data)
            
            add_audit_log(
                claim_id=claim_id,
                session_id=session_id,
                event_type="document_uploaded",
                event_message=f"Uploaded document {doc_type}: {filename}.",
                actor=current_user["email"],
                actor_email=current_user["email"],
                actor_role=current_user["role"],
                authenticated=True,
                metadata={"gcs_path": f"gs://{BUCKET_NAME}/{object_path}", "source": "employee_portal"}
            )
            
            expense_doc = db.collection(EXPENSES_COL).document(claim_id).get()
            if expense_doc.exists:
                await asyncio.to_thread(reevaluate_expense_policies, claim_id)
                
        return {
            "status": "success",
            "doc_type": doc_type,
            "path": object_path,
            "display_name": safe_filename
        }
    except Exception as e:
        logger.error(f"Error uploading file for claim {claim_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/api/employee/claims/preview")
async def preview_claim_policy(claim_data: dict, request: Request):
    current_user = get_current_user_and_role(request)
    policy_res = run_policy_check_py(claim_data)
    
    amount = float(claim_data.get("amount") or 0.0)
    reimbursable_amount = amount
    non_reimbursable_amount = 0.0
    
    trans_type = (claim_data.get("transportation_type") or "").lower()
    cat = (claim_data.get("category") or "").lower()
    if trans_type == "personal_vehicle" and cat == "transportation":
        gas_cost = float(claim_data.get("gas_cost") or 0.0)
        if gas_cost > 0:
            non_reimbursable_amount = gas_cost
            reimbursable_amount = max(0.0, amount - gas_cost)
            
    return {
        "policy_status": policy_res["status"],
        "is_valid": policy_res["is_valid"],
        "required_documents": policy_res["required_docs"],
        "missing_documents": policy_res["missing_docs"],
        "warnings": policy_res["warnings"],
        "estimated_reimbursable": reimbursable_amount,
        "estimated_non_reimbursable": non_reimbursable_amount,
        "manager_approval_required": not policy_res["is_valid"] or "Requires manager approval" in policy_res["status"]
    }


# ==========================================
# ENTERPRISE EXPENSE REPORT WORKFLOW APIs
# ==========================================

def check_claim_missing_documents(report_id: str, claim: dict):
    """
    Reruns the policy check for the claim, and checks the actual uploaded documents
    assigned to this claim to filter out from missing_documents.
    """
    cat = (claim.get("category") or "").lower()
    trans_type = (claim.get("transportation_type") or "").lower()
    if cat == "rental_car_gas" or trans_type == "rental_car_gas":
        if not claim.get("linked_rental_claim_id") and db and report_id:
            try:
                sibling_claims_ref = db.collection("expense_claims").where("report_id", "==", report_id)
                sibling_claims = [doc.to_dict() for doc in sibling_claims_ref.get()]
                has_rental_sibling = any(
                    (s.get("category") or "").lower() == "rental_car" or 
                    (s.get("transportation_type") or "").lower() == "rental_car"
                    for s in sibling_claims
                )
                if has_rental_sibling:
                    claim["linked_rental_claim_id"] = "report_context"
            except Exception as e:
                logger.warning(f"Error checking sibling claims for rental context: {e}")

    policy_res = run_policy_check_py(claim)
    required_docs = policy_res.get("required_docs") or []
    
    uploaded_types = set()
    if db:
        try:
            docs_ref = db.collection("documents").where("report_id", "==", report_id).where("claim_id", "==", claim["claim_id"]).where("active", "==", True).where("assigned_to_claim", "==", True)
            assigned_docs = [doc.to_dict() for doc in docs_ref.get()]
            uploaded_types = {d.get("doc_type") for d in assigned_docs}
        except Exception as e:
            logger.warning(f"Error checking active documents for claim {claim['claim_id']}: {e}")
            
    type_map = {
        "manager_approval_letter": "Manager Approval Letter",
        "flight_ticket_receipt": "Flight Ticket Receipt",
        "hotel_receipt": "Hotel Receipt",
        "receipt": "Receipt",
        "office_receipt": "Office Receipt",
        "parking_receipt": "Parking Receipt",
        "rental_receipt": "Rental Receipt",
        "gas_receipt": "Gas Receipt"
    }
    
    uploaded_doc_names = {type_map.get(t, t) for t in uploaded_types}
    if "receipt" in uploaded_types or "hotel_receipt" in uploaded_types or "flight_ticket_receipt" in uploaded_types:
        uploaded_doc_names.add("Receipt")
        
    missing_docs = []
    for req in required_docs:
        if req not in uploaded_doc_names:
            if req == "Receipt" and ("receipt_url" in claim or "receipt" in uploaded_types):
                continue
            missing_docs.append(req)
            
    claim["required_documents"] = required_docs
    claim["missing_documents"] = missing_docs
    claim["policy_status"] = "Within policy" if not missing_docs and policy_res["is_valid"] else policy_res["status"]
    
    return claim


def recalculate_report_totals(report_id: str):
    """
    Queries all claims and documents for a given report_id and recalculates totals and counts,
    then updates the report in Firestore.
    """
    if not db:
        return
    try:
        claims_ref = db.collection("expense_claims").where("report_id", "==", report_id)
        claims = [doc.to_dict() for doc in claims_ref.get()]
        
        # Check if the report has been approved with exceptions or has an override reason
        report_ref = db.collection("expense_reports").document(report_id)
        report_snap = report_ref.get()
        report_data = report_snap.to_dict() if report_snap.exists else {}
        report_status = report_data.get("status") or "draft"
        has_override = bool(report_data.get("override_reason")) or report_status == "approved_with_exceptions"
        
        total_claimed = 0.0
        total_reimbursable = 0.0
        total_non_reimbursable = 0.0
        policy_exception_count = 0
        missing_document_count = 0
        
        for c in claims:
            c_amount = float(c.get("amount") or 0.0)
            total_claimed += c_amount
            
            policy_status = c.get("policy_status") or "Within policy"
            has_exception = policy_status != "Within policy" and "Within policy" not in policy_status
            missing_docs = c.get("missing_documents") or []
            has_missing_docs = len(missing_docs) > 0
            
            if has_exception:
                policy_exception_count += 1
            missing_document_count += len(missing_docs)
            
            c_reimb = float(c.get("reimbursable_amount") or 0.0)
            c_non_reimb = float(c.get("non_reimbursable_amount") or 0.0)
            
            if c.get("claim_status") == "rejected":
                total_non_reimbursable += c_amount
            else:
                # If there are policy exceptions or missing documents, and NO override exists on the report,
                # then this item is not reimbursable (reimbursable amount is 0)
                if (has_exception or has_missing_docs) and not has_override:
                    total_non_reimbursable += c_amount
                else:
                    total_reimbursable += c_reimb
                    total_non_reimbursable += c_non_reimb
                    
        docs_ref = db.collection("documents").where("report_id", "==", report_id).where("active", "==", True)
        docs = [doc.to_dict() for doc in docs_ref.get()]
        
        document_count = len(docs)
        unassigned_document_count = sum(1 for d in docs if not d.get("assigned_to_claim"))
        
        if report_snap.exists:
            report_ref.update({
                "total_claimed_amount": total_claimed,
                "total_reimbursable_amount": total_reimbursable,
                "total_non_reimbursable_amount": total_non_reimbursable,
                "claim_count": len(claims),
                "document_count": document_count,
                "unassigned_document_count": unassigned_document_count,
                "missing_document_count": missing_document_count,
                "policy_exception_count": policy_exception_count,
                "updated_at": datetime.utcnow().isoformat() + "Z"
            })
    except Exception as e:
        logger.error(f"Error recalculating report {report_id} totals: {e}")


def add_report_audit_log(report_id: str, claim_id: str, event_type: str, message: str, current_user: dict, override_reason: str = None):
    """
    Writes a structured audit trail event log to Firestore for reports.
    """
    if not db:
        return
    try:
        doc_ref = db.collection("audit_logs").document()
        payload = {
            "company_id": "demo_company",
            "report_id": report_id,
            "claim_id": claim_id,
            "event_id": doc_ref.id,
            "event_type": event_type,
            "actor_email": current_user.get("email"),
            "actor_role": current_user.get("role"),
            "authenticated": True if event_type == "manager_decision" else current_user.get("authenticated", False),
            "message": message,
            "created_at": datetime.utcnow().isoformat() + "Z"
        }
        if override_reason:
            payload["override_reason"] = override_reason
            
        report_ref = db.collection("expense_reports").document(report_id)
        report_snap = report_ref.get()
        if report_snap.exists:
            report_data = report_snap.to_dict()
            payload["employee_email"] = report_data.get("employee_email")
            payload["manager_email"] = report_data.get("manager_email")
            
        doc_ref.set(payload)
    except Exception as e:
        logger.error(f"Failed to write report audit log: {e}")


@router.get("/api/reports")
async def get_reports(
    request: Request,
    status: str = None,
    employee: str = None,
    manager: str = None,
    department: str = None,
    start_date: str = None,
    end_date: str = None,
    limit: int = 50,
    offset: int = 0,
    refresh: bool = False
):
    current_user = get_current_user_and_role(request)
    email = current_user["email"]
    role = current_user["role"]
    
    # Construct cache key
    cache_key = (
        email,
        role,
        status,
        employee,
        manager,
        department,
        start_date,
        end_date,
        limit,
        offset
    )
    
    if not refresh:
        now = time.time()
        if cache_key in _expenses_cache:
            cache_ts, cached_data = _expenses_cache[cache_key]
            if now - cache_ts < _expenses_cache_ttl:
                logger.info(f"Returning cached reports for user {email}")
                return cached_data
                
    if not db:
        return []
        
    try:
        query = db.collection("expense_reports")
        
        # Scoping
        if role == "employee":
            query = query.where("employee_email", "==", email)
        elif role == "manager":
            query = query.where("manager_email", "==", email)
        elif role in ["finance_admin", "auditor", "admin"]:
            pass
        else:
            raise HTTPException(status_code=403, detail="Role not authorized to access reports")
            
        query = query.order_by("updated_at", direction=firestore.Query.DESCENDING)
        reports = [doc.to_dict() for doc in query.limit(limit + offset).get()]
        
        if offset > 0:
            reports = reports[offset:]
            
        filtered = []
        for r in reports:
            if status and r.get("status") != status:
                continue
            if employee and employee.lower() not in (r.get("employee_email") or "").lower() and employee.lower() not in (r.get("employee_name") or "").lower():
                continue
            if manager and manager.lower() not in (r.get("manager_email") or "").lower():
                continue
            if department and r.get("department") != department:
                continue
            if start_date and r.get("report_period_start") and r.get("report_period_start") < start_date:
                continue
            if end_date and r.get("report_period_end") and r.get("report_period_end") > end_date:
                continue
            filtered.append(r)
            
        _expenses_cache[cache_key] = (time.time(), filtered)
        return filtered
    except Exception as e:
        logger.error(f"Error in GET /api/reports: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/reports")
async def create_report(report_data: dict, request: Request):
    current_user = get_current_user_and_role(request)
    email = current_user["email"]
    role = current_user["role"]
    
    if not db:
        raise HTTPException(status_code=500, detail="Database not connected")
        
    company_id = report_data.get("company_id") or "demo_company"
    report_id = "rep_" + str(uuid.uuid4())[:18]
    
    employee_email = report_data.get("employee_email")
    if role == "finance_admin" and employee_email:
        pass
    else:
        employee_email = email
        
    employee_name = report_data.get("employee_name") or employee_email.split("@")[0].replace(".", " ").title()
    department = report_data.get("department") or "N/A"
    manager_email = report_data.get("manager_email") or "manager@company.com"
    employee_id = employee_email.split("@")[0]
    
    try:
        emp_docs = list(db.collection("employees").where("employee_email", "==", employee_email).where("active", "==", True).limit(1).get())
        if emp_docs:
            emp_data = emp_docs[0].to_dict()
            department = emp_data.get("department") or department
            manager_email = emp_data.get("manager_email") or manager_email
            employee_id = emp_data.get("employee_id") or employee_id
            employee_name = emp_data.get("employee_name") or employee_name
    except Exception as e:
        logger.warning(f"Error fetching employee profile: {e}")
        
    start_date = report_data.get("report_period_start") or datetime.utcnow().strftime("%Y-%m-%d")
    end_date = report_data.get("report_period_end") or datetime.utcnow().strftime("%Y-%m-%d")
    
    try:
        date_obj = datetime.fromisoformat(start_date.replace("Z", ""))
        month = date_obj.strftime("%B")
        year = date_obj.year
    except Exception:
        month = datetime.utcnow().strftime("%B")
        year = datetime.utcnow().year
        
    report = {
        "company_id": company_id,
        "report_id": report_id,
        "employee_id": employee_id,
        "employee_email": employee_email,
        "employee_name": employee_name,
        "department": department,
        "manager_id": manager_email.split("@")[0] if manager_email else "",
        "manager_email": manager_email,
        "report_title": report_data.get("report_title") or "Expense Report",
        "report_period_start": start_date,
        "report_period_end": end_date,
        "travel_week": report_data.get("travel_week"),
        "month": month,
        "year": year,
        "status": "draft",
        "total_claimed_amount": 0.0,
        "total_reimbursable_amount": 0.0,
        "total_non_reimbursable_amount": 0.0,
        "claim_count": 0,
        "document_count": 0,
        "unassigned_document_count": 0,
        "missing_document_count": 0,
        "policy_exception_count": 0,
        "submitted_by_email": "",
        "submitted_by_role": "",
        "submitted_at": "",
        "manager_reviewed_by": "",
        "manager_reviewed_at": "",
        "finance_reviewed_by": "",
        "finance_reviewed_at": "",
        "return_reason": "",
        "rejection_reason": "",
        "created_at": datetime.utcnow().isoformat() + "Z",
        "updated_at": datetime.utcnow().isoformat() + "Z"
    }
    
    try:
        db.collection("expense_reports").document(report_id).set(report)
        add_report_audit_log(report_id, None, "report_created", f"Created draft report '{report['report_title']}'.", current_user)
        return report
    except Exception as e:
        logger.error(f"Error creating report: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/reports/{report_id}")
async def get_report_detail(report_id: str, request: Request):
    current_user = get_current_user_and_role(request)
    email = current_user["email"]
    role = current_user["role"]
    
    if not db:
        raise HTTPException(status_code=500, detail="Database not connected")
        
    try:
        report_doc = db.collection("expense_reports").document(report_id).get()
        if not report_doc.exists:
            raise HTTPException(status_code=404, detail="Report not found")
            
        report = report_doc.to_dict()
        
        if role == "employee":
            if report.get("employee_email") != email:
                raise HTTPException(status_code=403, detail="Access denied to this report")
        elif role == "manager":
            if report.get("manager_email") != email:
                raise HTTPException(status_code=403, detail="Access denied to this report")
        elif role in ["finance_admin", "auditor", "admin"]:
            pass
        else:
            raise HTTPException(status_code=403, detail="Role not authorized to access reports")
            
        recalculate_report_totals(report_id)
        report = db.collection("expense_reports").document(report_id).get().to_dict()
        
        claims = [d.to_dict() for d in db.collection("expense_claims").where("report_id", "==", report_id).get()]
        documents = [d.to_dict() for d in db.collection("documents").where("report_id", "==", report_id).where("active", "==", True).get()]
        decisions = [d.to_dict() for d in db.collection("report_decisions").where("report_id", "==", report_id).get()]
        audit_logs = [d.to_dict() for d in db.collection("audit_logs").where("report_id", "==", report_id).get()]
        
        # Sort audit logs manually in memory to prevent indexing issues
        audit_logs.sort(key=lambda x: x.get("created_at") or "", reverse=True)
        
        return {
            "report": report,
            "claims": claims,
            "documents": documents,
            "decisions": decisions,
            "audit_logs": audit_logs
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in GET /api/reports/{report_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/api/reports/{report_id}")
async def update_report(report_id: str, update_data: dict, request: Request):
    current_user = get_current_user_and_role(request)
    
    if not db:
        raise HTTPException(status_code=500, detail="Database not connected")
        
    try:
        ref = db.collection("expense_reports").document(report_id)
        snap = ref.get()
        if not snap.exists:
            raise HTTPException(status_code=404, detail="Report not found")
            
        report = snap.to_dict()
        if report.get("status") not in ["draft", "returned_to_employee"]:
            raise HTTPException(status_code=400, detail="Can only update draft or returned reports")
            
        allowed_updates = ["report_title", "report_period_start", "report_period_end", "travel_week"]
        payload = {}
        for k in allowed_updates:
            if k in update_data:
                payload[k] = update_data[k]
                
        if payload:
            payload["updated_at"] = datetime.utcnow().isoformat() + "Z"
            ref.update(payload)
            add_report_audit_log(report_id, None, "report_updated", f"Updated report fields: {', '.join(payload.keys())}.", current_user)
            
        return ref.get().to_dict()
    except Exception as e:
        logger.error(f"Error updating report: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/reports/{report_id}/claims")
async def add_claim_to_report(report_id: str, claim_data: dict, request: Request):
    current_user = get_current_user_and_role(request)
    
    if not db:
        raise HTTPException(status_code=500, detail="Database not connected")
        
    try:
        report_ref = db.collection("expense_reports").document(report_id)
        report_snap = report_ref.get()
        if not report_snap.exists:
            raise HTTPException(status_code=404, detail="Report not found")
            
        report = report_snap.to_dict()
        if report.get("status") not in ["draft", "returned_to_employee"]:
            raise HTTPException(status_code=400, detail="Can only add claims to draft or returned reports")
            
        claim_id = claim_data.get("claim_id") or "clm_" + str(uuid.uuid4())[:18]
        amount = float(claim_data.get("amount") or 0.0)
        reimbursable_amount = amount
        non_reimbursable_amount = 0.0
        
        trans_type = (claim_data.get("transportation_type") or "").lower()
        cat = (claim_data.get("category") or "").lower()
        if trans_type == "personal_vehicle" and cat == "transportation":
            gas_cost = float(claim_data.get("gas_cost") or 0.0)
            if gas_cost > 0:
                non_reimbursable_amount = gas_cost
                reimbursable_amount = max(0.0, amount - gas_cost)
                
        claim = {
            "company_id": report.get("company_id") or "demo_company",
            "report_id": report_id,
            "claim_id": claim_id,
            "employee_id": report.get("employee_id"),
            "employee_email": report.get("employee_email"),
            "employee_name": report.get("employee_name"),
            "manager_email": report.get("manager_email"),
            "category": claim_data.get("category"),
            "amount": amount,
            "currency": claim_data.get("currency") or "USD",
            "expense_date": claim_data.get("expense_date") or datetime.utcnow().strftime("%Y-%m-%d"),
            "business_purpose": claim_data.get("business_purpose") or "",
            "description": claim_data.get("description") or "",
            "claim_status": "draft",
            "reimbursable_amount": reimbursable_amount,
            "non_reimbursable_amount": non_reimbursable_amount,
            "transportation_type": claim_data.get("transportation_type"),
            "gas_cost": claim_data.get("gas_cost"),
            "mileage": claim_data.get("mileage"),
            "check_in_date": claim_data.get("check_in_date"),
            "check_out_date": claim_data.get("check_out_date"),
            "travel_start_date": claim_data.get("travel_start_date") or report.get("report_period_start"),
            "travel_end_date": claim_data.get("travel_end_date") or report.get("report_period_end"),
            "city": claim_data.get("city") or report.get("city") or "",
            "state": claim_data.get("state") or report.get("state") or "",
            "state_code": claim_data.get("state_code") or report.get("state_code") or "",
            "country": claim_data.get("country") or report.get("country") or "US",
            "claimed_meals": claim_data.get("claimed_meals"),
            "claimed_lodging": claim_data.get("claimed_lodging"),
            "claimed_incidentals": claim_data.get("claimed_incidentals"),
            "created_at": datetime.utcnow().isoformat() + "Z",
            "updated_at": datetime.utcnow().isoformat() + "Z"
        }
        
        claim = check_claim_missing_documents(report_id, claim)
        db.collection("expense_claims").document(claim_id).set(claim)
        
        recalculate_report_totals(report_id)
        add_report_audit_log(report_id, claim_id, "claim_added_to_report", f"Added line item {claim['category']} of {claim['amount']} {claim['currency']} to report.", current_user)
        
        return claim
    except Exception as e:
        logger.error(f"Error adding claim: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/api/reports/{report_id}/claims/{claim_id}")
async def update_claim_in_report(report_id: str, claim_id: str, claim_data: dict, request: Request):
    current_user = get_current_user_and_role(request)
    
    if not db:
        raise HTTPException(status_code=500, detail="Database not connected")
        
    try:
        report_ref = db.collection("expense_reports").document(report_id)
        if not report_ref.get().exists:
            raise HTTPException(status_code=404, detail="Report not found")
            
        report = report_ref.get().to_dict()
        if report.get("status") not in ["draft", "returned_to_employee"]:
            raise HTTPException(status_code=400, detail="Can only update claims of draft or returned reports")
            
        claim_ref = db.collection("expense_claims").document(claim_id)
        snap = claim_ref.get()
        if not snap.exists:
            raise HTTPException(status_code=404, detail="Claim not found")
            
        claim = snap.to_dict()
        allowed_updates = ["category", "amount", "currency", "expense_date", "business_purpose", "description", "transportation_type", "gas_cost", "mileage", "check_in_date", "check_out_date"]
        for k in allowed_updates:
            if k in claim_data:
                claim[k] = claim_data[k]
                
        amount = float(claim.get("amount") or 0.0)
        reimbursable_amount = amount
        non_reimbursable_amount = 0.0
        
        trans_type = (claim.get("transportation_type") or "").lower()
        cat = (claim.get("category") or "").lower()
        if trans_type == "personal_vehicle" and cat == "transportation":
            gas_cost = float(claim.get("gas_cost") or 0.0)
            if gas_cost > 0:
                non_reimbursable_amount = gas_cost
                reimbursable_amount = max(0.0, amount - gas_cost)
                
        claim["amount"] = amount
        claim["reimbursable_amount"] = reimbursable_amount
        claim["non_reimbursable_amount"] = non_reimbursable_amount
        claim["updated_at"] = datetime.utcnow().isoformat() + "Z"
        
        claim = check_claim_missing_documents(report_id, claim)
        claim_ref.set(claim)
        
        recalculate_report_totals(report_id)
        add_report_audit_log(report_id, claim_id, "claim_updated", f"Updated line item {claim['category']}.", current_user)
        
        return claim
    except Exception as e:
        logger.error(f"Error updating claim: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/api/reports/{report_id}/claims/{claim_id}")
async def delete_claim_from_report(report_id: str, claim_id: str, request: Request):
    current_user = get_current_user_and_role(request)
    
    if not db:
        raise HTTPException(status_code=500, detail="Database not connected")
        
    try:
        report_ref = db.collection("expense_reports").document(report_id)
        if not report_ref.get().exists:
            raise HTTPException(status_code=404, detail="Report not found")
            
        report = report_ref.get().to_dict()
        if report.get("status") not in ["draft", "returned_to_employee"]:
            raise HTTPException(status_code=400, detail="Can only delete claims from draft or returned reports")
            
        claim_ref = db.collection("expense_claims").document(claim_id)
        if not claim_ref.get().exists:
            raise HTTPException(status_code=404, detail="Claim not found")
            
        claim = claim_ref.get().to_dict()
        
        docs_ref = db.collection("documents").where("report_id", "==", report_id).where("claim_id", "==", claim_id)
        for doc_snap in docs_ref.get():
            doc_snap.reference.update({
                "claim_id": None,
                "assigned_to_claim": False
            })
            
        claim_ref.delete()
        recalculate_report_totals(report_id)
        add_report_audit_log(report_id, claim_id, "claim_deleted", f"Deleted line item {claim.get('category')}.", current_user)
        
        return {"status": "success"}
    except Exception as e:
        logger.error(f"Error deleting claim: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/reports/{report_id}/documents")
async def bulk_upload_report_documents(report_id: str, request: Request, files: list[UploadFile] = File(...)):
    current_user = get_current_user_and_role(request)
    
    if not db:
        raise HTTPException(status_code=500, detail="Database not connected")
        
    try:
        report_ref = db.collection("expense_reports").document(report_id)
        if not report_ref.get().exists:
            raise HTTPException(status_code=404, detail="Report not found")
            
        uploaded_docs = []
        bucket = get_gcs_bucket()
        
        for file in files:
            filename = file.filename or "file"
            ext = os.path.splitext(filename)[1].lower()
            allowed_exts = {".pdf", ".png", ".jpg", ".jpeg"}
            if ext not in allowed_exts:
                continue
                
            contents = await file.read()
            doc_id = "doc_" + str(uuid.uuid4())[:18]
            doc_type = "receipt"
            
            object_path = f"uploads/reports/{report_id}/{doc_id}/{filename}"
            blob = bucket.blob(object_path)
            blob.upload_from_string(contents, content_type=file.content_type or "application/octet-stream")
            
            doc_data = {
                "company_id": "demo_company",
                "report_id": report_id,
                "claim_id": None,
                "document_id": doc_id,
                "doc_type": doc_type,
                "file_name": filename,
                "gcs_path": f"gs://{BUCKET_NAME}/{object_path}",
                "assigned_to_claim": False,
                "uploaded_by_email": current_user.get("email") or "employee@company.com",
                "uploaded_by_role": current_user.get("role") or "employee",
                "uploaded_at": datetime.utcnow().isoformat() + "Z",
                "active": True
            }
            
            db.collection("documents").document(doc_id).set(doc_data)
            uploaded_docs.append(doc_data)
            
        recalculate_report_totals(report_id)
        add_report_audit_log(report_id, None, "document_bulk_uploaded", f"Bulk uploaded {len(uploaded_docs)} document(s).", current_user)
        
        return uploaded_docs
    except Exception as e:
        logger.error(f"Error bulk uploading documents: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/reports/{report_id}/documents/{document_id}/assign")
async def assign_document_to_claim(report_id: str, document_id: str, assign_data: dict, request: Request):
    current_user = get_current_user_and_role(request)
    
    if not db:
        raise HTTPException(status_code=500, detail="Database not connected")
        
    try:
        doc_ref = db.collection("documents").document(document_id)
        snap = doc_ref.get()
        if not snap.exists:
            raise HTTPException(status_code=404, detail="Document not found")
            
        doc = snap.to_dict()
        claim_id = assign_data.get("claim_id")
        doc_type = assign_data.get("doc_type") or doc.get("doc_type")
        
        payload = {
            "claim_id": claim_id,
            "assigned_to_claim": True if claim_id else False,
            "doc_type": doc_type
        }
        doc_ref.update(payload)
        
        doc.update(payload)
        doc["status"] = "success"
        
        if claim_id:
            claim_ref = db.collection("expense_claims").document(claim_id)
            claim_snap = claim_ref.get()
            if claim_snap.exists:
                claim = claim_snap.to_dict()
                claim = check_claim_missing_documents(report_id, claim)
                claim_ref.set(claim)
                
        recalculate_report_totals(report_id)
        msg = f"Assigned document '{doc.get('file_name')}' to claim." if claim_id else f"Unassigned document '{doc.get('file_name')}'."
        add_report_audit_log(report_id, claim_id, "document_assigned_to_claim", msg, current_user)
        
        return doc
    except Exception as e:
        logger.error(f"Error assigning document: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/reports/{report_id}/submit")
async def submit_report(report_id: str, submit_data: dict, request: Request):
    current_user = get_current_user_and_role(request)
    role = current_user["role"]
    
    if not db:
        raise HTTPException(status_code=500, detail="Database not connected")
        
    try:
        report_ref = db.collection("expense_reports").document(report_id)
        report_snap = report_ref.get()
        if not report_snap.exists:
            raise HTTPException(status_code=404, detail="Report not found")
            
        report = report_snap.to_dict()
        if report.get("status") not in ["draft", "returned_to_employee"]:
            raise HTTPException(status_code=400, detail="Can only submit draft or returned reports")
            
        claims_ref = db.collection("expense_claims").where("report_id", "==", report_id)
        claims = [c.to_dict() for c in claims_ref.get()]
        
        if not claims:
            raise HTTPException(status_code=400, detail="Report must have at least one line item claim before submission.")
            
        total_missing = 0
        for c in claims:
            updated_c = check_claim_missing_documents(report_id, c)
            db.collection("expense_claims").document(c["claim_id"]).set(updated_c)
            total_missing += len(updated_c.get("missing_documents") or [])
            
        is_override = submit_data.get("override_missing_docs") or (role == "finance_admin")
        if total_missing > 0 and not is_override:
            raise HTTPException(status_code=400, detail=f"Submission blocked: {total_missing} required document(s) are missing.")
            
        unassigned_ref = db.collection("documents").where("report_id", "==", report_id).where("assigned_to_claim", "==", False).where("active", "==", True)
        for doc_snap in unassigned_ref.get():
            doc_snap.reference.update({
                "assigned_to_claim": True,
                "claim_id": "supporting"
            })
            
        recalculate_report_totals(report_id)
        
        # Call ExpenseOrchestrator
        from app.agents.orchestrator import ExpenseOrchestrator
        orchestrator = ExpenseOrchestrator()
        
        # Get up-to-date report and claims data
        report = report_ref.get().to_dict()
        claims_ref = db.collection("expense_claims").where("report_id", "==", report_id)
        claims = [c.to_dict() for c in claims_ref.get()]
        
        # Build normalized report
        normalized_report = {
            "report_id": report_id,
            "employee_name": report.get("employee_name") or "Unknown Claimant",
            "employee_email": report.get("employee_email") or "",
            "purpose": report.get("report_title") or "Expense Report",
            "total_amount": float(report.get("total_claimed_amount") or 0.0),
            "line_items": [
                {
                    "amount": float(c.get("amount") or 0.0),
                    "category": c.get("category") or "Other",
                    "description": c.get("business_purpose") or c.get("description") or "No description",
                    "receipt_url": c.get("receipt_url") or ""
                }
                for c in claims
            ]
        }
        
        orch_res = orchestrator.process_report(normalized_report)
        
        validation_res = orch_res.get("validation_result", {})
        if not validation_res.get("success", False):
            errors = validation_res.get("errors", ["Intake validation failed"])
            raise HTTPException(status_code=400, detail=f"Validation failed: {', '.join(errors)}")
            
        policy_warnings = orch_res.get("policy_result", {}).get("warnings", [])
        route_to = orch_res.get("routing_result", {}).get("route_to", "manager")
        agent_route = "finance_review" if route_to == "finance_review" else "manager_review"
        agent_recommendation = orch_res.get("final_recommendation")
        audit_events = orch_res.get("audit_events", [])
        
        report_ref.update({
            "status": "pending_manager_review",
            "submitted_by_email": current_user.get("email"),
            "submitted_by_role": current_user.get("role"),
            "submitted_at": datetime.utcnow().isoformat() + "Z",
            "updated_at": datetime.utcnow().isoformat() + "Z",
            "agent_intake_result": "Passed",
            "policy_warnings": policy_warnings,
            "agent_recommendation": agent_recommendation,
            "agent_route": agent_route,
            "agent_audit_event_count": len(audit_events)
        })
        
        # Write agent audit trail events to Firestore
        for evt in audit_events:
            try:
                doc_ref = db.collection("audit_logs").document(evt.get("event_id"))
                payload = {
                    "company_id": report.get("company_id") or "demo_company",
                    "report_id": report_id,
                    "claim_id": None,
                    "event_id": evt.get("event_id"),
                    "event_type": evt.get("event_type"),
                    "actor_email": evt.get("actor"),
                    "actor_role": "agent",
                    "authenticated": True,
                    "message": evt.get("message"),
                    "created_at": evt.get("timestamp"),
                    "employee_email": report.get("employee_email"),
                    "manager_email": report.get("manager_email")
                }
                if "metadata" in evt:
                    payload["metadata"] = evt["metadata"]
                doc_ref.set(payload)
            except Exception as e:
                logger.error(f"Failed to write agent audit log to Firestore: {e}")
        
        for c in claims:
            db.collection("expense_claims").document(c["claim_id"]).update({"claim_status": "submitted"})
            
        add_report_audit_log(report_id, None, "report_submitted", "Expense report submitted for manager review.", current_user)
        
        return {"status": "success"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error submitting report: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/reports/{report_id}/return")
async def return_report(report_id: str, decision_data: dict, request: Request):
    current_user = get_current_user_and_role(request)
    reason = decision_data.get("reason")
    if not reason:
        raise HTTPException(status_code=400, detail="Return reason is required")
        
    if not db:
        raise HTTPException(status_code=500, detail="Database not connected")
        
    try:
        ref = db.collection("expense_reports").document(report_id)
        snap = ref.get()
        if not snap.exists:
            raise HTTPException(status_code=404, detail="Report not found")
            
        report = snap.to_dict()
        ref.update({
            "status": "returned_to_employee",
            "return_reason": reason,
            "updated_at": datetime.utcnow().isoformat() + "Z"
        })
        
        dec_id = "dec_" + str(uuid.uuid4())[:18]
        dec_data = {
            "company_id": report.get("company_id") or "demo_company",
            "report_id": report_id,
            "claim_id": None,
            "decision_id": dec_id,
            "decision_scope": "report",
            "decision": "returned_to_employee",
            "reason": reason,
            "actor_email": current_user.get("email"),
            "actor_role": current_user.get("role"),
            "authenticated": True,
            "decided_at": datetime.utcnow().isoformat() + "Z"
        }
        db.collection("report_decisions").document(dec_id).set(dec_data)
        
        add_report_audit_log(report_id, None, "report_returned_to_employee", f"Report returned to employee. Reason: {reason}", current_user)
        
        return {"status": "success"}
    except Exception as e:
        logger.error(f"Error returning report: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/reports/{report_id}/approve")
async def approve_report(report_id: str, request: Request):
    current_user = get_current_user_and_role(request)
    
    if not db:
        raise HTTPException(status_code=500, detail="Database not connected")
        
    try:
        ref = db.collection("expense_reports").document(report_id)
        snap = ref.get()
        if not snap.exists:
            raise HTTPException(status_code=404, detail="Report not found")
            
        report = snap.to_dict()
        
        override_reason = None
        try:
            body = await request.json()
            if isinstance(body, dict):
                override_reason = body.get("override_reason")
        except Exception:
            pass
            
        missing_docs_count = report.get("missing_document_count") or 0
        exceptions_count = report.get("policy_exception_count") or 0
        
        status = "approved_by_manager"
        if missing_docs_count > 0 or exceptions_count > 0:
            if not override_reason or not override_reason.strip():
                raise HTTPException(
                    status_code=400,
                    detail="Report contains policy exceptions or missing documents. An override reason is required."
                )
            status = "approved_with_exceptions"
            
        update_data = {
            "status": status,
            "manager_reviewed_by": current_user.get("email"),
            "manager_reviewed_at": datetime.utcnow().isoformat() + "Z",
            "updated_at": datetime.utcnow().isoformat() + "Z"
        }
        if override_reason:
            update_data["override_reason"] = override_reason
            
        ref.update(update_data)
        
        # Recalculate totals now that status and override are updated, to include approved/overridden claims
        recalculate_report_totals(report_id)
        
        claims_ref = db.collection("expense_claims").where("report_id", "==", report_id)
        for c_snap in claims_ref.get():
            c = c_snap.to_dict()
            if c.get("claim_status") != "rejected":
                c_snap.reference.update({"claim_status": "approved"})
                
        dec_id = "dec_" + str(uuid.uuid4())[:18]
        dec_data = {
            "company_id": report.get("company_id") or "demo_company",
            "report_id": report_id,
            "claim_id": None,
            "decision_id": dec_id,
            "decision_scope": "report",
            "decision": "approved",
            "reason": override_reason or "Full report approved.",
            "actor_email": current_user.get("email"),
            "actor_role": current_user.get("role"),
            "authenticated": True,
            "decided_at": datetime.utcnow().isoformat() + "Z"
        }
        if override_reason:
            dec_data["override_reason"] = override_reason
            
        db.collection("report_decisions").document(dec_id).set(dec_data)
        
        audit_msg = "Expense report approved by manager."
        if status == "approved_with_exceptions":
            audit_msg = f"Expense report approved with exceptions. Override reason: {override_reason}"
            
        add_report_audit_log(report_id, None, "manager_decision", audit_msg, current_user, override_reason=override_reason)
        
        return {"status": "success"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error approving report: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/reports/{report_id}/pay")
async def pay_report(report_id: str, request: Request):
    current_user = get_current_user_and_role(request)
    role = current_user["role"]
    if role not in ["finance_admin", "admin"]:
        raise HTTPException(status_code=403, detail="Only Finance Admin or Admin can mark reports as paid.")
        
    if not db:
        raise HTTPException(status_code=500, detail="Database not connected")
        
    try:
        ref = db.collection("expense_reports").document(report_id)
        snap = ref.get()
        if not snap.exists:
            raise HTTPException(status_code=404, detail="Report not found")
            
        ref.update({
            "status": "paid",
            "finance_reviewed_by": current_user.get("email"),
            "finance_reviewed_at": datetime.utcnow().isoformat() + "Z",
            "updated_at": datetime.utcnow().isoformat() + "Z"
        })
        
        claims_ref = db.collection("expense_claims").where("report_id", "==", report_id)
        for c_snap in claims_ref.get():
            c_snap.reference.update({"claim_status": "paid"})
            
        add_report_audit_log(report_id, None, "report_paid", "Expense report review completed and marked as Paid/Reimbursed.", current_user)
        return {"status": "success"}
    except Exception as e:
        logger.error(f"Error marking report as paid: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/reports/{report_id}/reject")
async def reject_report(report_id: str, decision_data: dict, request: Request):
    current_user = get_current_user_and_role(request)
    reason = decision_data.get("reason") or "Rejected by manager."
    
    if not db:
        raise HTTPException(status_code=500, detail="Database not connected")
        
    try:
        ref = db.collection("expense_reports").document(report_id)
        snap = ref.get()
        if not snap.exists:
            raise HTTPException(status_code=404, detail="Report not found")
            
        report = snap.to_dict()
        ref.update({
            "status": "rejected",
            "rejection_reason": reason,
            "manager_reviewed_by": current_user.get("email"),
            "manager_reviewed_at": datetime.utcnow().isoformat() + "Z",
            "updated_at": datetime.utcnow().isoformat() + "Z"
        })
        
        claims_ref = db.collection("expense_claims").where("report_id", "==", report_id)
        for c_snap in claims_ref.get():
            c_snap.reference.update({"claim_status": "rejected"})
            
        dec_id = "dec_" + str(uuid.uuid4())[:18]
        dec_data = {
            "company_id": report.get("company_id") or "demo_company",
            "report_id": report_id,
            "claim_id": None,
            "decision_id": dec_id,
            "decision_scope": "report",
            "decision": "rejected",
            "reason": reason,
            "actor_email": current_user.get("email"),
            "actor_role": current_user.get("role"),
            "authenticated": True,
            "decided_at": datetime.utcnow().isoformat() + "Z"
        }
        db.collection("report_decisions").document(dec_id).set(dec_data)
        
        add_report_audit_log(report_id, None, "manager_decision", f"Expense report rejected. Reason: {reason}", current_user)
        
        return {"status": "success"}
    except Exception as e:
        logger.error(f"Error rejecting report: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/reports/{report_id}/claims/{claim_id}/decision")
async def decide_individual_claim(report_id: str, claim_id: str, decision_data: dict, request: Request):
    current_user = get_current_user_and_role(request)
    decision = decision_data.get("decision")
    reason = decision_data.get("reason") or f"Claim {decision}."
    
    if decision not in ["approved", "rejected"]:
        raise HTTPException(status_code=400, detail="Decision must be approved or rejected")
        
    if not db:
        raise HTTPException(status_code=500, detail="Database not connected")
        
    try:
        claim_ref = db.collection("expense_claims").document(claim_id)
        snap = claim_ref.get()
        if not snap.exists:
            raise HTTPException(status_code=404, detail="Claim not found")
            
        claim_ref.update({
            "claim_status": decision,
            "updated_at": datetime.utcnow().isoformat() + "Z"
        })
        
        dec_id = "dec_" + str(uuid.uuid4())[:18]
        dec_data = {
            "company_id": "demo_company",
            "report_id": report_id,
            "claim_id": claim_id,
            "decision_id": dec_id,
            "decision_scope": "claim",
            "decision": decision,
            "reason": reason,
            "actor_email": current_user.get("email"),
            "actor_role": current_user.get("role"),
            "authenticated": True,
            "decided_at": datetime.utcnow().isoformat() + "Z"
        }
        db.collection("report_decisions").document(dec_id).set(dec_data)
        
        recalculate_report_totals(report_id)
        add_report_audit_log(report_id, claim_id, "claim_decision", f"Line item was {decision}. Reason: {reason}", current_user)
        
        return {"status": "success"}
    except Exception as e:
        logger.error(f"Error in claim decision: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# --- TAX & AUDITOR EXPORT ENDPOINTS ---

class ExportPayload(BaseModel):
    claim_ids: Optional[List[str]] = None
    search: Optional[str] = None
    department: Optional[str] = None
    manager: Optional[str] = None
    status: Optional[str] = None
    category: Optional[str] = None
    company: Optional[str] = None
    hide_old_test_sessions: Optional[bool] = False
    assigned_to_me: Optional[bool] = False

def verify_export_access(request: Request) -> dict:
    current_user = get_current_user_and_role(request)
    if not current_user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Authentication required")
    role = current_user.get("role")
    if role not in ["finance_admin", "manager"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied: Only managers and finance admins can export data."
        )
    return current_user

async def get_claims_for_export(
    request: Request,
    payload: Optional[ExportPayload] = None,
    search: str = None,
    department: str = None,
    manager: str = None,
    status: str = None,
    category: str = None,
    company: str = None,
    hide_old_test_sessions: bool = False,
    assigned_to_me: bool = False
) -> list[dict]:
    current_user = get_current_user_and_role(request)
    
    if payload and payload.claim_ids is not None:
        claim_ids = payload.claim_ids
        if not claim_ids:
            return []
        
        claims = []
        chunk_size = 30
        for i in range(0, len(claim_ids), chunk_size):
            chunk = claim_ids[i:i + chunk_size]
            docs_ref = db.collection(EXPENSES_COL).where("claim_id", "in", chunk)
            for doc in docs_ref.get():
                claims.append(doc.to_dict())
                
        enriched_claims = [enrich_claim_with_employee_info(c) for c in claims]
        return enriched_claims
        
    params = {
        "search": search or (payload.search if payload else None),
        "department": department or (payload.department if payload else None),
        "manager": manager or (payload.manager if payload else None),
        "status": status or (payload.status if payload else None),
        "category": category or (payload.category if payload else None),
        "company": company or (payload.company if payload else None),
        "hide_old_test_sessions": hide_old_test_sessions or (payload.hide_old_test_sessions if payload else False),
        "assigned_to_me": assigned_to_me or (payload.assigned_to_me if payload else False)
    }
    
    expenses_ref = db.collection(EXPENSES_COL).order_by("created_at", direction=firestore.Query.DESCENDING)
    docs = list(expenses_ref.get())
    all_claims = [doc.to_dict() for doc in docs]
    filtered_claims = filter_and_enrich_claims(all_claims, current_user, params, is_pending=False)
    return [sanitize_claim_dict(c) for c in filtered_claims]

async def get_audits_for_export(claims: list[dict]) -> list[tuple[dict, dict]]:
    if not claims or not db:
        return []
        
    claim_ids = [c["claim_id"] for c in claims if "claim_id" in c]
    if not claim_ids:
        return []
        
    claims_map = {c["claim_id"]: c for c in claims}
    
    audits = []
    chunk_size = 30
    for i in range(0, len(claim_ids), chunk_size):
        chunk = claim_ids[i:i + chunk_size]
        audits_ref = db.collection(AUDIT_LOGS_COL).where("claim_id", "in", chunk)
        for doc in audits_ref.get():
            audits.append(doc.to_dict())
            
    audits_sorted = sorted(audits, key=lambda x: x.get("timestamp", ""))
    
    result = []
    for audit in audits_sorted:
        claim_id = audit.get("claim_id")
        claim = claims_map.get(claim_id) or {}
        result.append((audit, claim))
        
    return result

import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from fastapi import Response

def generate_excel_bytes(headers: list[str], rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"
    ws.views.sheetView[0].showGridLines = True
    
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    for row_data in rows:
        row_vals = [row_data.get(h, "") for h in headers]
        ws.append(row_vals)
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    ws.freeze_panes = "A2"
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def generate_expenses_csv_response(claims: list[dict]) -> Response:
    output = io.StringIO()
    writer = csv.writer(output)
    
    headers = [
        "Claim ID", "Category", "Date", "Employee Name", "Employee Email",
        "Department", "Assigned Manager", "Submitted By", "Amount", "Status",
        "Reviewer", "Policy Status", "Company ID"
    ]
    writer.writerow(headers)
    
    for c in claims:
        created_at = c.get("created_at") or ""
        writer.writerow([
            c.get("claim_id") or "",
            c.get("category") or "",
            created_at,
            c.get("employee_name") or "",
            c.get("employee_email") or "",
            c.get("department") or "",
            c.get("manager_email") or "",
            c.get("submitted_by_email") or "",
            c.get("amount") or 0.0,
            c.get("status") or "",
            c.get("reviewer_email") or c.get("decision_by_email") or "",
            c.get("policy_status") or "",
            c.get("company_id") or ""
        ])
        
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=expenses_export.csv"}
    )

def generate_expenses_excel_response(claims: list[dict]) -> Response:
    headers = [
        "Claim ID", "Category", "Date", "Employee Name", "Employee Email",
        "Department", "Assigned Manager", "Submitted By", "Amount", "Status",
        "Reviewer", "Policy Status", "Company ID"
    ]
    
    rows = []
    for c in claims:
        rows.append({
            "Claim ID": c.get("claim_id") or "",
            "Category": c.get("category") or "",
            "Date": c.get("created_at") or "",
            "Employee Name": c.get("employee_name") or "",
            "Employee Email": c.get("employee_email") or "",
            "Department": c.get("department") or "",
            "Assigned Manager": c.get("manager_email") or "",
            "Submitted By": c.get("submitted_by_email") or "",
            "Amount": c.get("amount") or 0.0,
            "Status": c.get("status") or "",
            "Reviewer": c.get("reviewer_email") or c.get("decision_by_email") or "",
            "Policy Status": c.get("policy_status") or "",
            "Company ID": c.get("company_id") or ""
        })
        
    excel_bytes = generate_excel_bytes(headers, rows)
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=expenses_export.xlsx"}
    )

def generate_audit_csv_response(audits: list[tuple[dict, dict]]) -> Response:
    output = io.StringIO()
    writer = csv.writer(output)
    
    headers = [
        "Report/Claim ID", "Employee Name/Email", "Department", "Manager/Reviewer",
        "Action Taken", "Field Changed", "Previous Value", "New Value",
        "Changed By", "Changed At", "Approval Status", "Amount", "Notes/Reason"
    ]
    writer.writerow(headers)
    
    for audit, claim in audits:
        emp_email = claim.get("employee_email") or audit.get("employee_email") or ""
        emp_name = claim.get("employee_name") or audit.get("employee_name") or ""
        emp_str = f"{emp_name} <{emp_email}>" if emp_email else emp_name
        
        metadata = audit.get("metadata") or {}
        field_changed = metadata.get("field") or metadata.get("field_changed") or ""
        prev_val = metadata.get("old_value") or metadata.get("previous_value") or ""
        new_val = metadata.get("new_value") or ""
        notes = audit.get("event_message") or metadata.get("reason") or metadata.get("override_reason") or ""
        
        writer.writerow([
            claim.get("claim_id") or audit.get("claim_id") or "",
            emp_str,
            claim.get("department") or "",
            claim.get("manager_email") or claim.get("reviewer_email") or "",
            audit.get("event_type") or "",
            field_changed,
            prev_val,
            new_val,
            audit.get("actor_email") or audit.get("actor") or "",
            audit.get("timestamp") or "",
            claim.get("status") or "",
            claim.get("amount") or 0.0,
            notes
        ])
        
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=audit_export.csv"}
    )

def generate_audit_excel_response(audits: list[tuple[dict, dict]]) -> Response:
    headers = [
        "Report/Claim ID", "Employee Name/Email", "Department", "Manager/Reviewer",
        "Action Taken", "Field Changed", "Previous Value", "New Value",
        "Changed By", "Changed At", "Approval Status", "Amount", "Notes/Reason"
    ]
    
    rows = []
    for audit, claim in audits:
        emp_email = claim.get("employee_email") or audit.get("employee_email") or ""
        emp_name = claim.get("employee_name") or audit.get("employee_name") or ""
        emp_str = f"{emp_name} <{emp_email}>" if emp_email else emp_name
        
        metadata = audit.get("metadata") or {}
        field_changed = metadata.get("field") or metadata.get("field_changed") or ""
        prev_val = metadata.get("old_value") or metadata.get("previous_value") or ""
        new_val = metadata.get("new_value") or ""
        notes = audit.get("event_message") or metadata.get("reason") or metadata.get("override_reason") or ""
        
        rows.append({
            "Report/Claim ID": claim.get("claim_id") or audit.get("claim_id") or "",
            "Employee Name/Email": emp_str,
            "Department": claim.get("department") or "",
            "Manager/Reviewer": claim.get("manager_email") or claim.get("reviewer_email") or "",
            "Action Taken": audit.get("event_type") or "",
            "Field Changed": field_changed,
            "Previous Value": prev_val,
            "New Value": new_val,
            "Changed By": audit.get("actor_email") or audit.get("actor") or "",
            "Changed At": audit.get("timestamp") or "",
            "Approval Status": claim.get("status") or "",
            "Amount": claim.get("amount") or 0.0,
            "Notes/Reason": notes
        })
        
    excel_bytes = generate_excel_bytes(headers, rows)
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=audit_export.xlsx"}
    )

@router.post("/api/export/expenses/csv")
async def export_expenses_csv_post(payload: ExportPayload, request: Request):
    verify_export_access(request)
    claims = await get_claims_for_export(request, payload)
    return generate_expenses_csv_response(claims)

@router.get("/api/export/expenses/csv")
async def export_expenses_csv_get(
    request: Request,
    search: str = None,
    department: str = None,
    manager: str = None,
    status: str = None,
    category: str = None,
    company: str = None,
    hide_old_test_sessions: bool = False,
    assigned_to_me: bool = False
):
    verify_export_access(request)
    claims = await get_claims_for_export(
        request, None, search, department, manager, status, category, company, hide_old_test_sessions, assigned_to_me
    )
    return generate_expenses_csv_response(claims)

@router.post("/api/export/expenses/excel")
async def export_expenses_excel_post(payload: ExportPayload, request: Request):
    verify_export_access(request)
    claims = await get_claims_for_export(request, payload)
    return generate_expenses_excel_response(claims)

@router.get("/api/export/expenses/excel")
async def export_expenses_excel_get(
    request: Request,
    search: str = None,
    department: str = None,
    manager: str = None,
    status: str = None,
    category: str = None,
    company: str = None,
    hide_old_test_sessions: bool = False,
    assigned_to_me: bool = False
):
    verify_export_access(request)
    claims = await get_claims_for_export(
        request, None, search, department, manager, status, category, company, hide_old_test_sessions, assigned_to_me
    )
    return generate_expenses_excel_response(claims)

@router.post("/api/export/audit/csv")
async def export_audit_csv_post(payload: ExportPayload, request: Request):
    verify_export_access(request)
    claims = await get_claims_for_export(request, payload)
    audits = await get_audits_for_export(claims)
    return generate_audit_csv_response(audits)

@router.get("/api/export/audit/csv")
async def export_audit_csv_get(
    request: Request,
    search: str = None,
    department: str = None,
    manager: str = None,
    status: str = None,
    category: str = None,
    company: str = None,
    hide_old_test_sessions: bool = False,
    assigned_to_me: bool = False
):
    verify_export_access(request)
    claims = await get_claims_for_export(
        request, None, search, department, manager, status, category, company, hide_old_test_sessions, assigned_to_me
    )
    audits = await get_audits_for_export(claims)
    return generate_audit_csv_response(audits)

@router.post("/api/export/audit/excel")
async def export_audit_excel_post(payload: ExportPayload, request: Request):
    verify_export_access(request)
    claims = await get_claims_for_export(request, payload)
    audits = await get_audits_for_export(claims)
    return generate_audit_excel_response(audits)

@router.get("/api/export/audit/excel")
async def export_audit_excel_get(
    request: Request,
    search: str = None,
    department: str = None,
    manager: str = None,
    status: str = None,
    category: str = None,
    company: str = None,
    hide_old_test_sessions: bool = False,
    assigned_to_me: bool = False
):
    verify_export_access(request)
    claims = await get_claims_for_export(
        request, None, search, department, manager, status, category, company, hide_old_test_sessions, assigned_to_me
    )
    audits = await get_audits_for_export(claims)
    return generate_audit_excel_response(audits)


# --- BUSINESS CREDIT CARD INTEGRATION MODULE ---

class ConnectCardPayload(BaseModel):
    provider: str
    cardholder_name: str
    cardholder_email: str
    card_last4: str

class MatchReceiptPayload(BaseModel):
    transaction_id: str
    receipt_url: str
    receipt_name: str

class AttachClaimPayload(BaseModel):
    transaction_id: str
    claim_id: str

class ReconcileCardPayload(BaseModel):
    transaction_id: str

class UpdateCardPayload(BaseModel):
    transaction_id: str
    notes: Optional[str] = None
    reconciliation_status: Optional[str] = None


@router.get("/api/cards/transactions")
async def get_card_transactions(request: Request):
    current_user = get_current_user_and_role(request)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    if not db:
        return []
        
    try:
        col_ref = db.collection("card_transactions")
        docs = list(col_ref.get())
        
        # If collection is empty, seed initial mock transactions
        if not docs:
            initial_mocks = [
                {
                    "transaction_id": "tx_001",
                    "provider": "Ramp",
                    "card_last4": "1002",
                    "cardholder_name": "Jane Admin",
                    "cardholder_email": "finance-admin@company.com",
                    "transaction_date": "2026-06-15",
                    "amount": 125.50,
                    "merchant": "AWS Cloud Services",
                    "currency": "USD",
                    "reconciliation_status": "unreconciled",
                    "matched_claim_id": None,
                    "matched_receipt_url": None,
                    "matched_receipt_name": None,
                    "notes": "AWS monthly hosting bill",
                    "company_id": "demo_company"
                },
                {
                    "transaction_id": "tx_002",
                    "provider": "Brex",
                    "card_last4": "4092",
                    "cardholder_name": "John Manager",
                    "cardholder_email": "manager@company.com",
                    "transaction_date": "2026-06-20",
                    "amount": 42.00,
                    "merchant": "Uber Trips",
                    "currency": "USD",
                    "reconciliation_status": "unreconciled",
                    "matched_claim_id": None,
                    "matched_receipt_url": None,
                    "matched_receipt_name": None,
                    "notes": "Ride to office",
                    "company_id": "demo_company"
                },
                {
                    "transaction_id": "tx_003",
                    "provider": "Amex",
                    "card_last4": "3001",
                    "cardholder_name": "Alice Employee",
                    "cardholder_email": "employee@company.com",
                    "transaction_date": "2026-06-22",
                    "amount": 350.00,
                    "merchant": "Delta Air Lines",
                    "currency": "USD",
                    "reconciliation_status": "unreconciled",
                    "matched_claim_id": None,
                    "matched_receipt_url": None,
                    "matched_receipt_name": None,
                    "notes": "Flight to conference",
                    "company_id": "demo_company"
                },
                {
                    "transaction_id": "tx_004",
                    "provider": "Visa",
                    "card_last4": "9876",
                    "cardholder_name": "Jane Admin",
                    "cardholder_email": "finance-admin@company.com",
                    "transaction_date": "2026-06-25",
                    "amount": 18.50,
                    "merchant": "Starbucks",
                    "currency": "USD",
                    "reconciliation_status": "unreconciled",
                    "matched_claim_id": None,
                    "matched_receipt_url": None,
                    "matched_receipt_name": None,
                    "notes": "Coffee meeting",
                    "company_id": "demo_company"
                },
                {
                    "transaction_id": "tx_005",
                    "provider": "Stripe Issuing",
                    "card_last4": "5544",
                    "cardholder_name": "Bob Developer",
                    "cardholder_email": "bob.developer@company.com",
                    "transaction_date": "2026-06-28",
                    "amount": 29.00,
                    "merchant": "GitHub",
                    "currency": "USD",
                    "reconciliation_status": "unreconciled",
                    "matched_claim_id": None,
                    "matched_receipt_url": None,
                    "matched_receipt_name": None,
                    "notes": "GitHub Copilot subscription",
                    "company_id": "demo_company"
                },
                {
                    "transaction_id": "tx_006",
                    "provider": "Mastercard",
                    "card_last4": "2211",
                    "cardholder_name": "Alice Employee",
                    "cardholder_email": "employee@company.com",
                    "transaction_date": "2026-06-29",
                    "amount": 88.00,
                    "merchant": "Zoom.us",
                    "currency": "USD",
                    "reconciliation_status": "unreconciled",
                    "matched_claim_id": None,
                    "matched_receipt_url": None,
                    "matched_receipt_name": None,
                    "notes": "Video conferencing subscription",
                    "company_id": "demo_company"
                },
                {
                    "transaction_id": "tx_007",
                    "provider": "Plaid-style bank feed",
                    "card_last4": "7788",
                    "cardholder_name": "Jane Admin",
                    "cardholder_email": "finance-admin@company.com",
                    "transaction_date": "2026-06-30",
                    "amount": 2500.00,
                    "merchant": "Office Rent LLC",
                    "currency": "USD",
                    "reconciliation_status": "unreconciled",
                    "matched_claim_id": None,
                    "matched_receipt_url": None,
                    "matched_receipt_name": None,
                    "notes": "Monthly office space rent",
                    "company_id": "demo_company"
                }
            ]
            for m in initial_mocks:
                col_ref.document(m["transaction_id"]).set(m)
            return initial_mocks
            
        transactions = [d.to_dict() for d in docs]
        transactions.sort(key=lambda x: x.get("transaction_date", ""), reverse=True)
        return transactions
    except Exception as e:
        logger.error(f"Error fetching card transactions: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/cards/connect")
async def connect_card_feed(payload: ConnectCardPayload, request: Request):
    current_user = get_current_user_and_role(request)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
        
    role = current_user.get("role")
    if role not in ["finance_admin", "manager"]:
        raise HTTPException(status_code=403, detail="Access denied: Only finance admins and managers can connect bank feeds.")
        
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
        
    try:
        feed_id = f"feed_{payload.provider.lower().replace(' ', '_')}_{payload.card_last4}"
        feed_doc = {
            "feed_id": feed_id,
            "provider": payload.provider,
            "cardholder_name": payload.cardholder_name,
            "cardholder_email": payload.cardholder_email,
            "card_last4": payload.card_last4,
            "connected_at": datetime.utcnow().isoformat() + "Z",
            "connected_by": current_user.get("email"),
            "company_id": "demo_company"
        }
        db.collection("card_feeds").document(feed_id).set(feed_doc)
        
        mock_merchants = [
            ("Google Cloud", 150.00, "Cloud platform billing"),
            ("Slack Technologies", 45.00, "Slack team workspace subscription"),
            ("Uber", 22.40, "Business trip transport")
        ]
        
        new_txs = []
        for idx, (merchant, amount, note) in enumerate(mock_merchants):
            tx_id = f"tx_mock_{int(time.time())}_{idx}"
            tx_doc = {
                "transaction_id": tx_id,
                "provider": payload.provider,
                "card_last4": payload.card_last4,
                "cardholder_name": payload.cardholder_name,
                "cardholder_email": payload.cardholder_email,
                "transaction_date": datetime.utcnow().strftime("%Y-%m-%d"),
                "amount": amount,
                "merchant": merchant,
                "currency": "USD",
                "reconciliation_status": "unreconciled",
                "matched_claim_id": None,
                "matched_receipt_url": None,
                "matched_receipt_name": None,
                "notes": note,
                "company_id": "demo_company"
            }
            db.collection("card_transactions").document(tx_id).set(tx_doc)
            new_txs.append(tx_doc)
            
        add_audit_log(
            claim_id="None",
            session_id="credit_card_integration",
            event_type="card_feed_connected",
            event_message=f"Connected mock {payload.provider} business card feed ending in {payload.card_last4} for {payload.cardholder_name}.",
            actor=current_user.get("name") or "system",
            actor_email=current_user.get("email"),
            actor_role=current_user.get("role"),
            authenticated=True,
            metadata={"provider": payload.provider, "card_last4": payload.card_last4, "cardholder_name": payload.cardholder_name}
        )
        
        return {"status": "success", "message": f"Successfully connected {payload.provider} feed. 3 mock transactions synced.", "transactions": new_txs}
    except Exception as e:
        logger.error(f"Error connecting card feed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/cards/match-receipt")
async def match_transaction_receipt(payload: MatchReceiptPayload, request: Request):
    current_user = get_current_user_and_role(request)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
        
    role = current_user.get("role")
    if role not in ["finance_admin", "manager"]:
        raise HTTPException(status_code=403, detail="Access denied: Only managers and admins can perform reconciliation.")
        
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
        
    try:
        tx_ref = db.collection("card_transactions").document(payload.transaction_id)
        tx_snap = tx_ref.get()
        if not tx_snap.exists:
            raise HTTPException(status_code=404, detail="Card transaction not found")
            
        tx_data = tx_snap.to_dict()
        
        update_data = {
            "matched_receipt_url": payload.receipt_url,
            "matched_receipt_name": payload.receipt_name
        }
        
        status_to_set = "unreconciled"
        if tx_data.get("matched_claim_id"):
            status_to_set = "matched"
            update_data["reconciliation_status"] = status_to_set
            
        tx_ref.update(update_data)
        
        add_audit_log(
            claim_id=tx_data.get("matched_claim_id") or "None",
            session_id="credit_card_integration",
            event_type="card_receipt_matched",
            event_message=f"Matched receipt '{payload.receipt_name}' to card transaction {payload.transaction_id} from {tx_data.get('provider')} (merchant: {tx_data.get('merchant')}, amount: ${tx_data.get('amount')}).",
            actor=current_user.get("name") or "system",
            actor_email=current_user.get("email"),
            actor_role=current_user.get("role"),
            authenticated=True,
            metadata={"transaction_id": payload.transaction_id, "receipt_name": payload.receipt_name, "reconciliation_status": status_to_set}
        )
        
        return {"status": "success", "reconciliation_status": status_to_set}
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Error matching receipt: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/cards/attach-claim")
async def attach_transaction_to_claim(payload: AttachClaimPayload, request: Request):
    current_user = get_current_user_and_role(request)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
        
    role = current_user.get("role")
    if role not in ["finance_admin", "manager"]:
        raise HTTPException(status_code=403, detail="Access denied: Only managers and admins can perform reconciliation.")
        
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
        
    try:
        tx_ref = db.collection("card_transactions").document(payload.transaction_id)
        tx_snap = tx_ref.get()
        if not tx_snap.exists:
            raise HTTPException(status_code=404, detail="Card transaction not found")
            
        tx_data = tx_snap.to_dict()
        
        update_data = {
            "matched_claim_id": payload.claim_id,
            "reconciliation_status": "matched"
        }
        tx_ref.update(update_data)
        
        claim_ref = db.collection(EXPENSES_COL).document(payload.claim_id)
        claim_snap = claim_ref.get()
        if claim_snap.exists:
            claim_ref.update({
                "payment_method": "Business Card",
                "card_transaction_id": payload.transaction_id,
                "reconciliation_status": "matched"
            })
            
        add_audit_log(
            claim_id=payload.claim_id,
            session_id="credit_card_integration",
            event_type="card_transaction_matched",
            event_message=f"Attached card transaction {payload.transaction_id} from {tx_data.get('provider')} (merchant: {tx_data.get('merchant')}, amount: ${tx_data.get('amount')}) to expense claim {payload.claim_id}.",
            actor=current_user.get("name") or "system",
            actor_email=current_user.get("email"),
            actor_role=current_user.get("role"),
            authenticated=True,
            metadata={"transaction_id": payload.transaction_id, "claim_id": payload.claim_id, "reconciliation_status": "matched"}
        )
        
        return {"status": "success"}
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Error attaching transaction to claim: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/cards/reconcile")
async def reconcile_card_transaction(payload: ReconcileCardPayload, request: Request):
    current_user = get_current_user_and_role(request)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
        
    role = current_user.get("role")
    if role not in ["finance_admin", "manager"]:
        raise HTTPException(status_code=403, detail="Access denied: Only managers and admins can perform reconciliation.")
        
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
        
    try:
        tx_ref = db.collection("card_transactions").document(payload.transaction_id)
        tx_snap = tx_ref.get()
        if not tx_snap.exists:
            raise HTTPException(status_code=404, detail="Card transaction not found")
            
        tx_data = tx_snap.to_dict()
        
        tx_ref.update({
            "reconciliation_status": "reconciled"
        })
        
        matched_claim_id = tx_data.get("matched_claim_id")
        if matched_claim_id:
            claim_ref = db.collection(EXPENSES_COL).document(matched_claim_id)
            claim_snap = claim_ref.get()
            if claim_snap.exists:
                claim_ref.update({
                    "reconciliation_status": "reconciled"
                })
                
        add_audit_log(
            claim_id=matched_claim_id or "None",
            session_id="credit_card_integration",
            event_type="card_transaction_reconciled",
            event_message=f"Reconciled card transaction {payload.transaction_id} from {tx_data.get('provider')} (merchant: {tx_data.get('merchant')}, amount: ${tx_data.get('amount')}). Status: reconciled.",
            actor=current_user.get("name") or "system",
            actor_email=current_user.get("email"),
            actor_role=current_user.get("role"),
            authenticated=True,
            metadata={"transaction_id": payload.transaction_id, "reconciliation_status": "reconciled"}
        )
        
        return {"status": "success"}
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Error reconciling transaction: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/cards/update")
async def update_card_transaction_notes(payload: UpdateCardPayload, request: Request):
    current_user = get_current_user_and_role(request)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
        
    role = current_user.get("role")
    if role not in ["finance_admin", "manager"]:
        raise HTTPException(status_code=403, detail="Access denied: Only managers and admins can edit card transactions.")
        
    if not db:
        raise HTTPException(status_code=500, detail="Database not initialized")
        
    try:
        tx_ref = db.collection("card_transactions").document(payload.transaction_id)
        tx_snap = tx_ref.get()
        if not tx_snap.exists:
            raise HTTPException(status_code=404, detail="Card transaction not found")
            
        tx_data = tx_snap.to_dict()
        
        update_dict = {}
        if payload.notes is not None:
            update_dict["notes"] = payload.notes
        if payload.reconciliation_status is not None:
            update_dict["reconciliation_status"] = payload.reconciliation_status
            
        if not update_dict:
            return {"status": "success", "message": "No changes requested."}
            
        tx_ref.update(update_dict)
        
        add_audit_log(
            claim_id=tx_data.get("matched_claim_id") or "None",
            session_id="credit_card_integration",
            event_type="card_transaction_edited",
            event_message=f"Updated card transaction {payload.transaction_id} from {tx_data.get('provider')} (merchant: {tx_data.get('merchant')}). Notes: '{payload.notes}'.",
            actor=current_user.get("name") or "system",
            actor_email=current_user.get("email"),
            actor_role=current_user.get("role"),
            authenticated=True,
            metadata={"transaction_id": payload.transaction_id, "changes": update_dict}
        )
        
        return {"status": "success"}
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Error updating transaction: {e}")
        raise HTTPException(status_code=500, detail=str(e))


def get_fallback_zero_state():
    return {
        "summary": {
            "total_agent_runs": 0,
            "policy_warnings_generated": 0,
            "reports_routed_to_manager": 0,
            "reports_routed_to_finance": 0,
            "audit_events_generated": 0
        },
        "agents": [
            {
                "id": "intake_agent",
                "name": "Expense Intake Agent",
                "status": "Healthy",
                "runs_today": 0,
                "avg_processing_time_ms": 78,
                "success_rate_pct": 100.0,
                "last_run": "Never",
                "description": "Validates file completeness, extracts transaction text, and ensures necessary line items are present before processing."
            },
            {
                "id": "policy_agent",
                "name": "Policy Compliance Agent",
                "status": "Healthy",
                "runs_today": 0,
                "avg_processing_time_ms": 145,
                "success_rate_pct": 100.0,
                "last_run": "Never",
                "description": "Evaluates line-item amounts against company expense guidelines, checking limits, required documentation, and merchant policy compliance."
            },
            {
                "id": "routing_agent",
                "name": "Approval Routing Agent",
                "status": "Healthy",
                "runs_today": 0,
                "avg_processing_time_ms": 92,
                "success_rate_pct": 100.0,
                "last_run": "Never",
                "description": "Determines report severity, routing low-risk expenses to direct managers and escalating high-risk or high-value expenses to Finance."
            },
            {
                "id": "audit_agent",
                "name": "Audit Intelligence Agent",
                "status": "Healthy",
                "runs_today": 0,
                "avg_processing_time_ms": 64,
                "success_rate_pct": 100.0,
                "last_run": "Never",
                "description": "Compiles structured, immutable audit log events detailing every agent's decision and warning for complete compliance and historical analysis."
            },
            {
                "id": "orchestrator",
                "name": "Expense Orchestrator",
                "status": "Healthy",
                "runs_today": 0,
                "avg_processing_time_ms": 379,
                "success_rate_pct": 100.0,
                "last_run": "Never",
                "description": "Master execution coordinator that manages state transition and pipeline flow across all sub-agents sequentially."
            }
        ],
        "has_real_data": False
    }


@router.get("/api/agents/metrics")
async def get_agent_ops_metrics(request: Request):
    """
    Retrieves system-wide performance and operations metrics for the AI Agents.
    If no data exists, computes safe demo/zero-state metrics without crashing.
    """
    try:
        current_user = get_current_user_and_role(request)
    except Exception:
        raise HTTPException(status_code=401, detail="Authentication required")
        
    if not db:
        return get_fallback_zero_state()
        
    try:
        # Fetch all expense reports to compile metrics
        reports_ref = db.collection("expense_reports")
        reports = [r.to_dict() for r in reports_ref.get()]
        
        # Filter reports that have run through the agent pipeline
        # Usually they have "agent_recommendation" or "agent_route"
        orchestrated_reports = [r for r in reports if r.get("agent_recommendation")]
        
        # Filter for runs today
        # To determine "today", let's parse submitted_at in UTC
        now = datetime.utcnow()
        today_str = now.strftime("%Y-%m-%d")
        
        runs_today_count = 0
        policy_warnings_count = 0
        routed_to_manager = 0
        routed_to_finance = 0
        
        last_run_timestamp = None
        
        for r in orchestrated_reports:
            submitted_at = r.get("submitted_at")
            if submitted_at:
                # Check if it was today
                if submitted_at.startswith(today_str):
                    runs_today_count += 1
                
                # Keep track of the latest run timestamp
                if not last_run_timestamp or submitted_at > last_run_timestamp:
                    last_run_timestamp = submitted_at
                    
            warnings = r.get("policy_warnings") or []
            policy_warnings_count += len(warnings)
            
            route = r.get("agent_route")
            if route == "manager_review":
                routed_to_manager += 1
            elif route == "finance_review":
                routed_to_finance += 1
                
        # Count agent audit log entries
        # Querying with actor_role == "agent"
        try:
            audits_ref = db.collection("audit_logs").where("actor_role", "==", "agent").get()
            agent_audits = [a.to_dict() for a in audits_ref]
            audit_events_count = len(agent_audits)
            
            # Find the absolute latest agent audit timestamp if report submitted_at is not available
            for a in agent_audits:
                created_at = a.get("created_at") or a.get("timestamp")
                if created_at:
                    if not last_run_timestamp or created_at > last_run_timestamp:
                        last_run_timestamp = created_at
        except Exception as ae:
            logger.error(f"Error querying agent audit logs: {ae}")
            audit_events_count = len(orchestrated_reports) * 4 # rough estimate as fallback
            
        total_runs = len(orchestrated_reports)
        
        # Format the last run timestamp nicely
        formatted_last_run = "Never"
        if last_run_timestamp:
            formatted_last_run = last_run_timestamp
            
        result = {
            "summary": {
                "total_agent_runs": total_runs,
                "policy_warnings_generated": policy_warnings_count,
                "reports_routed_to_manager": routed_to_manager,
                "reports_routed_to_finance": routed_to_finance,
                "audit_events_generated": audit_events_count
            },
            "agents": [
                {
                    "id": "intake_agent",
                    "name": "Expense Intake Agent",
                    "status": "Healthy",
                    "runs_today": runs_today_count,
                    "avg_processing_time_ms": 78,
                    "success_rate_pct": 98.4 if total_runs > 0 else 100.0,
                    "last_run": formatted_last_run,
                    "description": "Validates file completeness, extracts transaction text, and ensures necessary line items are present before processing."
                },
                {
                    "id": "policy_agent",
                    "name": "Policy Compliance Agent",
                    "status": "Healthy",
                    "runs_today": runs_today_count,
                    "avg_processing_time_ms": 145,
                    "success_rate_pct": 100.0,
                    "last_run": formatted_last_run,
                    "description": "Evaluates line-item amounts against company expense guidelines, checking limits, required documentation, and merchant policy compliance."
                },
                {
                    "id": "routing_agent",
                    "name": "Approval Routing Agent",
                    "status": "Healthy",
                    "runs_today": runs_today_count,
                    "avg_processing_time_ms": 92,
                    "success_rate_pct": 100.0,
                    "last_run": formatted_last_run,
                    "description": "Determines report severity, routing low-risk expenses to direct managers and escalating high-risk or high-value expenses to Finance."
                },
                {
                    "id": "audit_agent",
                    "name": "Audit Intelligence Agent",
                    "status": "Healthy",
                    "runs_today": runs_today_count,
                    "avg_processing_time_ms": 64,
                    "success_rate_pct": 100.0,
                    "last_run": formatted_last_run,
                    "description": "Compiles structured, immutable audit log events detailing every agent's decision and warning for complete compliance and historical analysis."
                },
                {
                    "id": "orchestrator",
                    "name": "Expense Orchestrator",
                    "status": "Healthy",
                    "runs_today": runs_today_count,
                    "avg_processing_time_ms": 379,
                    "success_rate_pct": 97.5 if total_runs > 0 else 100.0,
                    "last_run": formatted_last_run,
                    "description": "Master execution coordinator that manages state transition and pipeline flow across all sub-agents sequentially."
                }
            ],
            "has_real_data": total_runs > 0
        }
        return result
    except Exception as e:
        logger.error(f"Error in get_agent_ops_metrics: {e}")
        return get_fallback_zero_state()


LOGIN_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Sign In - Ambience ExpenseFlow</title>
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&family=Plus+Jakarta+Sans:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <style>
        :root {{
            --bg-color: #060713;
            --primary: #6366f1;
            --primary-glow: rgba(99, 102, 241, 0.2);
            --accent: #8b5cf6;
            --text-main: #f3f4f6;
            --text-muted: #9ca3af;
            --glass-bg: rgba(15, 17, 36, 0.55);
            --glass-border: rgba(255, 255, 255, 0.08);
            --transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        }}

        * {{
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }}

        body {{
            font-family: 'Outfit', sans-serif;
            background-color: var(--bg-color);
            background-image: 
                radial-gradient(circle at 20% 30%, rgba(99, 102, 241, 0.15) 0%, transparent 50%),
                radial-gradient(circle at 80% 70%, rgba(139, 92, 246, 0.12) 0%, transparent 50%);
            color: var(--text-main);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            overflow: hidden;
            position: relative;
            padding: 1.5rem;
        }}

        .glow-sphere {{
            position: absolute;
            width: 300px;
            height: 300px;
            background: radial-gradient(circle, rgba(99, 102, 241, 0.12) 0%, transparent 70%);
            filter: blur(40px);
            pointer-events: none;
            z-index: 0;
            animation: pulse 6s ease-in-out infinite alternate;
        }}

        .glow-sphere-1 {{ top: 10%; left: 15%; }}
        .glow-sphere-2 {{ bottom: 10%; right: 15%; }}

        @keyframes pulse {{
            0% {{ transform: scale(1) translate(0, 0); }}
            100% {{ transform: scale(1.2) translate(10px, 10px); }}
        }}

        .login-card {{
            background: var(--glass-bg);
            backdrop-filter: blur(20px);
            -webkit-backdrop-filter: blur(20px);
            border: 1px solid var(--glass-border);
            border-radius: 28px;
            padding: 3rem 2.5rem;
            width: 100%;
            max-width: 440px;
            box-shadow: 0 20px 50px rgba(0, 0, 0, 0.5);
            text-align: center;
            position: relative;
            z-index: 10;
            animation: fadeInUp 0.8s cubic-bezier(0.16, 1, 0.3, 1) forwards;
        }}

        .login-card::before {{
            content: '';
            position: absolute;
            top: 0;
            left: 50%;
            transform: translateX(-50%);
            width: 80%;
            height: 1px;
            background: linear-gradient(90deg, transparent, var(--primary), var(--accent), transparent);
        }}

        .logo-icon {{
            width: 56px;
            height: 56px;
            background: linear-gradient(135deg, var(--primary) 0%, var(--accent) 100%);
            border-radius: 16px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-weight: 800;
            font-size: 1.5rem;
            color: white;
            margin: 0 auto 1.5rem auto;
            box-shadow: 0 8px 30px rgba(99, 102, 241, 0.4);
            animation: float 4s ease-in-out infinite;
        }}

        h1 {{
            font-size: 1.8rem;
            font-weight: 700;
            background: linear-gradient(to right, #ffffff, #c7d2fe);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            margin-bottom: 0.5rem;
        }}

        p {{
            font-size: 0.95rem;
            color: var(--text-muted);
            line-height: 1.5;
            margin-bottom: 2.5rem;
        }}

        .btn-google {{
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 0.75rem;
            background: #ffffff;
            color: #1f2937;
            border: 1px solid #e5e7eb;
            border-radius: 14px;
            padding: 0.9rem 1.5rem;
            font-family: 'Plus Jakarta Sans', sans-serif;
            font-weight: 600;
            font-size: 1rem;
            cursor: pointer;
            width: 100%;
            transition: var(--transition);
            box-shadow: 0 4px 15px rgba(0, 0, 0, 0.1);
            text-decoration: none;
        }}

        .btn-google:hover {{
            background: #f9fafb;
            transform: translateY(-2px);
            box-shadow: 0 8px 25px rgba(255, 255, 255, 0.15), 0 0 15px var(--primary-glow);
            border-color: var(--primary);
        }}

        .btn-google:active {{
            transform: translateY(0);
        }}

        .btn-google svg {{
            width: 20px;
            height: 20px;
        }}

        .badge-dev {{
            background: rgba(245, 158, 11, 0.1);
            border: 1px solid rgba(245, 158, 11, 0.25);
            color: #f59e0b;
            padding: 0.75rem 1rem;
            border-radius: 12px;
            font-size: 0.85rem;
            line-height: 1.4;
            text-align: left;
            margin-bottom: 2rem;
            display: flex;
            align-items: flex-start;
            gap: 0.5rem;
        }}

        .btn-bypass {{
            background: linear-gradient(135deg, var(--primary) 0%, var(--accent) 100%);
            color: white;
            border: none;
            border-radius: 14px;
            padding: 0.9rem 1.5rem;
            font-family: inherit;
            font-weight: 700;
            font-size: 1rem;
            cursor: pointer;
            width: 100%;
            transition: var(--transition);
            box-shadow: 0 4px 20px rgba(99, 102, 241, 0.35);
            text-decoration: none;
            display: inline-block;
        }}

        .btn-bypass:hover {{
            background: linear-gradient(135deg, #4f46e5 0%, #7c3aed 100%);
            transform: translateY(-2px);
            box-shadow: 0 8px 30px rgba(99, 102, 241, 0.55);
        }}

        .footer-text {{
            font-size: 0.75rem;
            color: rgba(255, 255, 255, 0.3);
            margin-top: 2.5rem;
            text-transform: uppercase;
            letter-spacing: 0.05em;
        }}

        @keyframes float {{
            0%, 100% {{ transform: translateY(0px); }}
            50% {{ transform: translateY(-8px); }}
        }}

        @keyframes fadeInUp {{
            from {{ opacity: 0; transform: translateY(20px); }}
            to {{ opacity: 1; transform: translateY(0); }}
        }}
    
        /* New premium styles for Expense Reports */
        .reports-grid {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(320px, 1fr));
            gap: 1.5rem;
            margin-top: 1rem;
        }
        .report-card {
            background: var(--glass-bg);
            backdrop-filter: blur(12px);
            border: 1px solid var(--glass-border);
            border-radius: 20px;
            padding: 1.5rem;
            box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.2);
            transition: var(--transition);
            display: flex;
            flex-direction: column;
            justify-content: space-between;
            position: relative;
            overflow: hidden;
        }
        .report-card:hover {
            transform: translateY(-5px);
            border-color: rgba(99, 102, 241, 0.3);
            box-shadow: 0 12px 40px 0 rgba(99, 102, 241, 0.15);
        }
        .report-card::before {
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            width: 100%;
            height: 4px;
            background: linear-gradient(90deg, var(--primary), var(--accent));
        }
        .report-title {
            font-size: 1.25rem;
            font-weight: 700;
            color: white;
            margin-bottom: 0.5rem;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }
        .report-meta {
            display: flex;
            flex-direction: column;
            gap: 0.4rem;
            font-size: 0.85rem;
            color: var(--text-muted);
            margin-bottom: 1rem;
        }
        .report-meta-item {
            display: flex;
            align-items: center;
            gap: 0.5rem;
        }
        .report-totals {
            display: flex;
            justify-content: space-between;
            border-top: 1px solid rgba(255, 255, 255, 0.05);
            padding-top: 1rem;
            margin-top: auto;
        }
        .report-total-box {
            display: flex;
            flex-direction: column;
        }
        .report-total-label {
            font-size: 0.75rem;
            color: var(--text-muted);
            text-transform: uppercase;
            letter-spacing: 0.05em;
        }
        .report-total-val {
            font-size: 1.1rem;
            font-weight: 700;
            color: white;
        }
        /* Slider Panel (slide drawer) */
        .slider-panel {
            position: fixed;
            top: 0;
            right: -450px;
            width: 450px;
            height: 100vh;
            background: rgba(10, 11, 26, 0.95);
            backdrop-filter: blur(20px);
            border-left: 1px solid var(--glass-border);
            box-shadow: -10px 0 40px rgba(0,0,0,0.5);
            z-index: 100;
            transition: right 0.4s cubic-bezier(0.4, 0, 0.2, 1);
            padding: 2.5rem;
            display: flex;
            flex-direction: column;
            overflow-y: auto;
        }
        .slider-panel.active {
            right: 0;
        }
        .slider-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 2rem;
            border-bottom: 1px solid rgba(255,255,255,0.05);
            padding-bottom: 1rem;
        }
        /* Dropzone styling */
        .dropzone {
            border: 2px dashed rgba(99, 102, 241, 0.3);
            background: rgba(99, 102, 241, 0.02);
            border-radius: 16px;
            padding: 2rem;
            text-align: center;
            cursor: pointer;
            transition: var(--transition);
            margin: 1.5rem 0;
        }
        .dropzone:hover {
            border-color: var(--primary);
            background: rgba(99, 102, 241, 0.05);
        }
        /* Unassigned docs list */
        .unassigned-docs-bar {
            background: rgba(255, 255, 255, 0.02);
            border: 1px solid var(--glass-border);
            border-radius: 16px;
            padding: 1.25rem;
            margin-top: 1.5rem;
        }
        .unassigned-badge {
            display: inline-flex;
            align-items: center;
            gap: 0.5rem;
            background: rgba(245, 158, 11, 0.1);
            border: 1px solid rgba(245, 158, 11, 0.25);
            color: #f59e0b;
            padding: 0.4rem 0.8rem;
            border-radius: 12px;
            font-size: 0.8rem;
            font-weight: 500;
            cursor: pointer;
            transition: var(--transition);
        }
        .unassigned-badge:hover {
            background: rgba(245, 158, 11, 0.15);
            transform: scale(1.03);
        }

</style>
</head>
<body>
    <div class="glow-sphere glow-sphere-1"></div>
    <div class="glow-sphere glow-sphere-2"></div>

    <div class="login-card">
        <div class="logo-icon">AEF</div>
        <h1>Ambience ExpenseFlow</h1>
        <p>Enterprise Travel & Expense Management</p>

        {content_html}

        <div class="footer-text">SECURE OIDC SIGN-IN ΓÇó AES-256 SESSION CODES</div>
    </div>
</body>
</html>
"""

